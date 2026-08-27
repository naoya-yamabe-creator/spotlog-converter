import streamlit as st
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side
from copy import copy
import io
import re
import os
import time
import base64
from datetime import datetime, timezone, timedelta

# 画面基本設定
st.set_page_config(
    page_title="spotlog ➔ 基準様式 Excel自動変換ツール",
    page_icon="📄",
    layout="centered"
)

# カスタムCSS
st.markdown("""
    <link rel="stylesheet" href="https://fonts.googleapis.com/css2?family=Material+Symbols+Outlined:opsz,wght,FILL,GRAD@20..48,100..700,0..1,-50..200" />
    <style>
    .material-symbols-outlined {
        font-family: 'Material Symbols Outlined' !important;
        font-weight: normal !important;
        font-style: normal !important;
        font-size: 20px;
        line-height: 1 !important;
        letter-spacing: normal !important;
        text-transform: none !important;
        display: inline-flex !important;
        align-items: center !important;
        justify-content: center !important;
        vertical-align: middle !important;
        white-space: nowrap !important;
        word-wrap: normal !important;
        direction: ltr !important;
        -webkit-font-smoothing: antialiased !important;
    }
    /* アプリ全体をライトモード（純白背景）で完全固定 */
    html, body, [data-testid="stAppViewContainer"], .stApp {
        background-color: #FFFFFF !important;
        color: #1E293B !important;
    }
    /* Streamlit標準のDeployボタン・右上の三本点メニュー・ヘッダー・フッターを完全非表示 */
    #MainMenu,
    header,
    header[data-testid="stHeader"],
    div[data-testid="stToolbar"],
    div[data-testid="stDecoration"],
    div[data-testid="stStatusWidget"],
    footer,
    .stDeployButton {
        display: none !important;
        visibility: hidden !important;
        height: 0 !important;
    }
    /* ページ全体コンテナを均等な左右マージンで完全中央配置 */
    .block-container {
        max-width: 720px !important;
        padding-top: 3.0rem !important;
        padding-left: 1.5rem !important;
        padding-right: 1.5rem !important;
        margin-left: auto !important;
        margin-right: auto !important;
    }
    .main-header {
        font-size: clamp(1.25rem, 3.8vw, 2.2rem) !important;
        font-weight: 800;
        color: #1E293B;
        text-align: center;
        margin-bottom: 0.75rem !important;
    }
    .sub-header {
        font-size: clamp(0.85rem, 2.2vw, 1.0rem) !important;
        color: #64748B;
        text-align: center;
        line-height: 1.7 !important;
        margin-bottom: 2.2rem !important;
    }
    /* ドロップゾーン見出し行：単一HTML行による100%完璧な水平中央同期 */
    .dropzone-header-flex {
        display: flex !important;
        align-items: center !important;
        justify-content: space-between !important;
        width: 100% !important;
        margin-bottom: 0.75rem !important;
        position: relative !important;
    }
    .dropzone-title-text {
        font-size: 0.95rem !important;
        font-weight: 600 !important;
        color: #1E293B !important;
        line-height: 1 !important;
        display: flex !important;
        align-items: center !important;
    }
    .spec-popover {
        position: relative !important;
        display: inline-flex !important;
        align-items: center !important;
    }
    .spec-popover summary {
        list-style: none !important;
        outline: none !important;
        display: flex !important;
        align-items: center !important;
        justify-content: center !important;
    }
    .spec-popover summary::-webkit-details-marker {
        display: none !important;
    }
    .spec-info-btn {
        display: inline-flex !important;
        align-items: center !important;
        justify-content: center !important;
        width: 19px !important;
        height: 19px !important;
        min-width: 19px !important;
        min-height: 19px !important;
        border-radius: 50% !important;
        background-color: #64748B !important;
        color: #FFFFFF !important;
        font-size: 11.5px !important;
        font-weight: 700 !important;
        font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif !important;
        cursor: pointer !important;
        user-select: none !important;
        line-height: 1 !important;
        transition: background-color 0.2s ease, transform 0.2s ease !important;
    }
    .spec-info-btn:hover {
        background-color: #0A84FF !important;
        transform: scale(1.1) !important;
    }
    .spec-popover-card {
        position: absolute !important;
        right: 0 !important;
        top: calc(100% + 8px) !important;
        width: min(500px, 88vw) !important;
        background: #FFFFFF !important;
        border-radius: 14px !important;
        padding: 1.3rem 1.5rem !important;
        box-shadow: 0 14px 35px rgba(15, 23, 42, 0.18), 0 4px 10px rgba(15, 23, 42, 0.08) !important;
        z-index: 999999 !important;
        color: #1E293B !important;
        text-align: left !important;
        box-sizing: border-box !important;
    }
    .spec-popover-card h4 {
        margin: 0 0 0.9rem 0 !important;
        font-size: 1.05rem !important;
        font-weight: 700 !important;
        color: #0F172A !important;
    }
    .spec-popover-card ul {
        margin: 0 !important;
        padding: 0 !important;
        list-style: none !important;
        font-size: 0.88rem !important;
        line-height: 1.65 !important;
    }
    .spec-popover-card li {
        margin-bottom: 0.65rem !important;
        position: relative !important;
        padding-left: 1.15rem !important;
        line-height: 1.65 !important;
    }
    .spec-popover-card li::before {
        content: "" !important;
        position: absolute !important;
        left: 0.25rem !important;
        top: 0.55rem !important;
        width: 4.5px !important;
        height: 4.5px !important;
        border-radius: 50% !important;
        background-color: #1E293B !important;
    }
    .spec-popover-card li:last-child {
        margin-bottom: 0 !important;
    }
    .spec-popover-card code {
        background: #F1F5F9 !important;
        color: #1E293B !important;
        font-weight: 600 !important;
        padding: 0.15rem 0.45rem !important;
        border-radius: 6px !important;
        border: 1px solid #E2E8F0 !important;
        font-size: 0.84rem !important;
        word-break: break-all !important;
    }
    /* ドロップゾーン内の英語文言を日本語で直感的にわかりやすく表示 */
    [data-testid="stFileUploadDropzone"] button {
        font-size: 0 !important;
        padding: 0.55rem 1.2rem !important;
        display: inline-flex !important;
        align-items: center !important;
        justify-content: center !important;
    }
    [data-testid="stFileUploadDropzone"] button::after {
        content: "📁 ファイルを選択" !important;
        font-size: 0.92rem !important;
        font-weight: 600 !important;
        color: #1E293B !important;
    }
    [data-testid="stFileUploadDropzone"] small {
        font-size: 0 !important;
    }
    [data-testid="stFileUploadDropzone"] small::after {
        content: "または、ここにファイルをドラッグ＆ドロップ（.xlsx 形式 / 複数可）" !important;
        font-size: 0.85rem !important;
        color: #64748B !important;
        display: block !important;
        margin-top: 0.4rem !important;
    }
    /* ドロップゾーン下の余分な余白を削減 */
    div[data-testid="stFileUploader"] {
        margin-bottom: 0 !important;
        padding-bottom: 0 !important;
    }
    /* 変換完了サマリーカード（アプリ全体のブルー・ホワイト基調で完全統一） */
    .result-summary-card {
        background-color: #F8FAFC !important;
        border: 1.5px solid #BFDBFE !important;
        border-radius: 16px !important;
        padding: 1.4rem 1.6rem !important;
        margin-top: 1.2rem !important;
        box-shadow: 0 4px 16px rgba(10, 132, 255, 0.06) !important;
        color: #1E293B !important;
    }
    .result-header {
        display: flex !important;
        align-items: center !important;
        gap: 0.85rem !important;
        padding-bottom: 0.9rem !important;
        border-bottom: 1px solid #E2E8F0 !important;
        margin-bottom: 1rem !important;
    }
    .result-icon-badge {
        font-size: 1.6rem !important;
        line-height: 1 !important;
    }
    .result-main-title {
        font-size: 1.12rem !important;
        font-weight: 700 !important;
        color: #1E293B !important;
        line-height: 1.3 !important;
    }
    .result-sub-title {
        font-size: 0.85rem !important;
        color: #64748B !important;
        margin-top: 0.15rem !important;
    }
    .result-details-grid {
        display: flex !important;
        flex-direction: column !important;
        gap: 0.55rem !important;
        font-size: 0.92rem !important;
    }
    .result-detail-row {
        display: flex !important;
        align-items: center !important;
        flex-wrap: wrap !important;
        gap: 0.4rem !important;
    }
    .result-label {
        color: #475569 !important;
        font-weight: 600 !important;
    }
    .result-value {
        color: #1E293B !important;
    }
    .file-path-badge {
        background: #F1F5F9 !important;
        color: #1E293B !important;
        font-weight: 600 !important;
        padding: 0.2rem 0.5rem !important;
        border-radius: 6px !important;
        border: 1px solid #E2E8F0 !important;
        font-size: 0.88rem !important;
        word-break: break-all !important;
    }
    /* 変換失敗・全件スキップ時のエラーカード（赤い枠線＋清潔な黒文字） */
    .error-summary-card {
        background-color: #FEF2F2 !important;
        border: 1.5px solid #FECACA !important;
        border-radius: 14px !important;
        padding: 1.1rem 1.4rem !important;
        margin-top: 1.2rem !important;
        color: #1E293B !important;
        font-size: 0.95rem !important;
        display: flex !important;
        align-items: center !important;
        gap: 0.6rem !important;
    }
    .error-summary-title {
        font-weight: 700 !important;
        color: #1E293B !important;
    }
    .error-summary-sub {
        color: #64748B !important;
        font-size: 0.88rem !important;
    }
    .result-skipped-section {
        margin-top: 1rem !important;
        padding-top: 0.9rem !important;
        border-top: 1px solid #E2E8F0 !important;
    }
    .result-skipped-title {
        font-weight: 700 !important;
        color: #475569 !important;
        font-size: 0.9rem !important;
        margin-bottom: 0.45rem !important;
        display: flex !important;
        align-items: center !important;
    }
    .result-skipped-scroll {
        max-height: 180px !important;
        overflow-y: auto !important;
        background: #FFFFFF !important;
        border: 1px solid #E2E8F0 !important;
        border-radius: 8px !important;
        padding: 0.6rem 0.9rem !important;
    }
    .result-skipped-scroll ul {
        margin: 0 !important;
        padding-left: 1.1rem !important;
        font-size: 0.85rem !important;
        line-height: 1.6 !important;
        color: #475569 !important;
    }
    .result-skipped-scroll li {
        margin-bottom: 0.3rem !important;
    }
    .result-skipped-scroll li:last-child {
        margin-bottom: 0 !important;
    }
    /* 変換ボタン：ドロップゾーンの幅に100%完全連動（常に75%比率で真ん中に中央揃え） */
    [data-testid="stElementContainer"]:has([data-testid="stButton"]),
    div[data-testid="stButton"],
    .stButton {
        display: flex !important;
        justify-content: center !important;
        align-items: center !important;
        width: 100% !important;
        margin: -0.2rem auto 0.5rem auto !important;
        padding-top: 0 !important;
        text-align: center !important;
    }
    [data-testid="stButton"] > button,
    .stButton > button {
        width: 75% !important;
        margin-left: auto !important;
        margin-right: auto !important;
        display: flex !important;
        align-items: center !important;
        justify-content: center !important;
        background: #0A84FF !important;
        background-color: #0A84FF !important;
        color: white !important;
        font-weight: 700 !important;
        font-size: 1.1rem !important;
        padding: 0.85rem 1.5rem !important;
        border-radius: 12px !important;
        border: none !important;
        box-shadow: 0 4px 14px rgba(10, 132, 255, 0.35) !important;
        transition: all 0.25s ease !important;
    }
    [data-testid="stButton"] > button:hover,
    .stButton > button:hover {
        background: #3090FF !important;
        background-color: #3090FF !important;
        box-shadow: 0 6px 18px rgba(10, 132, 255, 0.45) !important;
        transform: translateY(-1px) !important;
    }
    /* ========================================================
       ファイルアップローダー（DropZoneスタイル完全一致）
       ======================================================== */
    div[data-testid="stFileUploader"] section,
    div[data-testid="stFileUploaderDropzone"],
    section[data-testid="stFileUploadDropzone"],
    [data-testid="stFileUploader"] > section,
    [data-testid="stFileUploadDropzone"] {
        border-radius: 16px !important;
        border: 2px dashed #94A3B8 !important;
        background-color: #F8FAFC !important;
        padding: 3.6rem 2rem !important;
        min-height: 250px !important;
        display: flex !important;
        flex-direction: column !important;
        align-items: center !important;
        justify-content: center !important;
        text-align: center !important;
        margin-bottom: 0 !important;
        transition: all 0.2s ease-in-out !important;
    }
    
    /* マウスホバー時：コードの #8E8E93 に合わせた上品なグレー変化 */
    div[data-testid="stFileUploader"] section:hover,
    div[data-testid="stFileUploaderDropzone"]:hover,
    section[data-testid="stFileUploadDropzone"]:hover,
    [data-testid="stFileUploadDropzone"]:hover {
        border-color: #8E8E93 !important;
        background-color: #F1F5F9 !important;
    }
    </style>
""", unsafe_allow_html=True)


def sanitize_sheet_name(name, max_len=31):
    """Excelのタブ名に利用できない文字を除去し、最大31文字に制限"""
    invalid_chars = r'[\\/*?:\[\]]'
    clean_name = re.sub(invalid_chars, '', str(name)).strip()
    if not clean_name:
        clean_name = "Sheet"
    return clean_name[:max_len]


def find_template_file(template_type, base_dir):
    """指定タイプのテンプレートファイルを検索"""
    if template_type == "鍼灸":
        candidates = ["基準様式　鍼.xlsx", "基準様式_鍼.xlsx", "基準様式鍼.xlsx"]
    else:
        candidates = ["基準様式　マ.xlsx", "基準様式_マッサージ.xlsx", "基準様式マ.xlsx", "基準様式_マ.xlsx"]
        
    for name in candidates:
        full_path = os.path.join(base_dir, name)
        if os.path.exists(full_path):
            return full_path
    return None


def format_currency_str(val):
    """数値を3桁カンマ区切りに変換（例: 4650 -> 4,650）"""
    if val is None or str(val).strip() == "":
        return ""
    try:
        num = int(round(float(str(val).replace(',', '').replace('円', '').strip())))
        return f"{num:,}"
    except Exception:
        return str(val)


def safe_apply_grid_row(ws, start_r, end_r, price_val='', count_val='', total_val='', price_val2=None, count_val2=None, total_val2=None, is_boxed=False):
    """
    金額計算行を正確な列セル[単価][円×][回数][回＝][金額][円]で配置
    - is_boxed=True: 通所用（標準格子枠・外枠罫線あり）
    - is_boxed=False: 通所以外（見本通り「円×」「回＝」を右寄りにシフト配置・罫線なし）
    """
    col_bf = openpyxl.utils.column_index_from_string('BF')
    col_dn = openpyxl.utils.column_index_from_string('DN')

    if is_boxed:
        # 通所用 (標準配置・枠線あり)
        c_p_start, c_p_end = openpyxl.utils.column_index_from_string('BF'), openpyxl.utils.column_index_from_string('BV')
        c_m1_start, c_m1_end = openpyxl.utils.column_index_from_string('BW'), openpyxl.utils.column_index_from_string('CC')
        c_c_start, c_c_end = openpyxl.utils.column_index_from_string('CD'), openpyxl.utils.column_index_from_string('CP')
        c_m2_start, c_m2_end = openpyxl.utils.column_index_from_string('CQ'), openpyxl.utils.column_index_from_string('CW')
        c_t_start, c_t_end = openpyxl.utils.column_index_from_string('CX'), openpyxl.utils.column_index_from_string('DJ')
        c_m3_start, c_m3_end = openpyxl.utils.column_index_from_string('DK'), openpyxl.utils.column_index_from_string('DN')
    else:
        # 通所以外（見本通り「円×」「回＝」を右寄りに配置し、金額欄CZ:DJを6桁が自然に収まる適正幅に調整）
        c_p_start, c_p_end = openpyxl.utils.column_index_from_string('BF'), openpyxl.utils.column_index_from_string('CD')
        c_m1_start, c_m1_end = openpyxl.utils.column_index_from_string('CE'), openpyxl.utils.column_index_from_string('CK')
        c_c_start, c_c_end = openpyxl.utils.column_index_from_string('CL'), openpyxl.utils.column_index_from_string('CQ')
        c_m2_start, c_m2_end = openpyxl.utils.column_index_from_string('CR'), openpyxl.utils.column_index_from_string('CY')
        c_t_start, c_t_end = openpyxl.utils.column_index_from_string('CZ'), openpyxl.utils.column_index_from_string('DJ')
        c_m3_start, c_m3_end = openpyxl.utils.column_index_from_string('DK'), openpyxl.utils.column_index_from_string('DN')

    # 1. 範囲内と交差する既存の結合をすべて安全に解除
    to_unmerge = []
    for rng in list(ws.merged_cells.ranges):
        min_c, min_r, max_c, max_r = rng.bounds
        if not (max_r < start_r or min_r > end_r or max_c < col_bf or min_c > col_dn):
            to_unmerge.append(str(rng))
            
    for rng_str in to_unmerge:
        try:
            ws.unmerge_cells(rng_str)
        except Exception:
            pass

    # 2. セル内容と罫線をリセット（外枠・行境界のみ保持、内部の縦線は完全にゼロ）
    thin = openpyxl.styles.Side(style='thin', color='000000')
    font_ms = openpyxl.styles.Font(name='ＭＳ 明朝', size=9.5)

    for r in range(start_r, end_r + 1):
        for c in range(col_bf, col_dn + 1):
            cell = ws.cell(row=r, column=c)
            cell.value = None
            t_s = thin if r == start_r else None
            b_s = thin if r == end_r else None
            l_s = thin if c == col_bf else None
            r_s = thin if c == col_dn else None
            cell.border = openpyxl.styles.Border(top=t_s, bottom=b_s, left=l_s, right=r_s)
            cell.font = font_ms

    has_two = price_val2 and str(price_val2).strip() not in ['', '0', '0円']
    
    p1_str = format_currency_str(price_val) if price_val else ''
    c1_str = str(count_val) if (count_val is not None and str(count_val).strip() != '') else ''
    t1_str = format_currency_str(total_val) if total_val else ''

    if has_two:
        mid_r = start_r + (end_r - start_r + 1) // 2
        p2_str = format_currency_str(price_val2) if price_val2 else ''
        c2_str = str(count_val2) if (count_val2 is not None and str(count_val2).strip() != '') else ''
        t2_str = format_currency_str(total_val2) if total_val2 else ''

        # 上段
        ws.merge_cells(start_row=start_r, start_column=c_p_start, end_row=mid_r-1, end_column=c_p_end)
        ws.merge_cells(start_row=start_r, start_column=c_c_start, end_row=mid_r-1, end_column=c_c_end)
        ws.merge_cells(start_row=start_r, start_column=c_t_start, end_row=mid_r-1, end_column=c_t_end)
        ws.cell(row=start_r, column=c_p_start, value=p1_str).alignment = openpyxl.styles.Alignment(horizontal='right', vertical='center')
        ws.cell(row=start_r, column=c_c_start, value=c1_str).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        ws.cell(row=start_r, column=c_t_start, value=t1_str).alignment = openpyxl.styles.Alignment(horizontal='right', vertical='center')

        # 下段
        ws.merge_cells(start_row=mid_r, start_column=c_p_start, end_row=end_r, end_column=c_p_end)
        ws.merge_cells(start_row=mid_r, start_column=c_c_start, end_row=end_r, end_column=c_c_end)
        ws.merge_cells(start_row=mid_r, start_column=c_t_start, end_row=end_r, end_column=c_t_end)
        ws.cell(row=mid_r, column=c_p_start, value=p2_str).alignment = openpyxl.styles.Alignment(horizontal='right', vertical='center')
        ws.cell(row=mid_r, column=c_c_start, value=c2_str).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        ws.cell(row=mid_r, column=c_t_start, value=t2_str).alignment = openpyxl.styles.Alignment(horizontal='right', vertical='center')
    else:
        # 1段（中央の横線なし）
        ws.merge_cells(start_row=start_r, start_column=c_p_start, end_row=end_r, end_column=c_p_end)
        ws.merge_cells(start_row=start_r, start_column=c_c_start, end_row=end_r, end_column=c_c_end)
        ws.merge_cells(start_row=start_r, start_column=c_t_start, end_row=end_r, end_column=c_t_end)
        ws.cell(row=start_r, column=c_p_start, value=p1_str).alignment = openpyxl.styles.Alignment(horizontal='right', vertical='center')
        ws.cell(row=start_r, column=c_c_start, value=c1_str).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        ws.cell(row=start_r, column=c_t_start, value=t1_str).alignment = openpyxl.styles.Alignment(horizontal='right', vertical='center')

    # 記号枠（円×, 回＝, 円）
    ws.merge_cells(start_row=start_r, start_column=c_m1_start, end_row=end_r, end_column=c_m1_end)
    ws.merge_cells(start_row=start_r, start_column=c_m2_start, end_row=end_r, end_column=c_m2_end)
    ws.merge_cells(start_row=start_r, start_column=c_m3_start, end_row=end_r, end_column=c_m3_end)
    ws.cell(row=start_r, column=c_m1_start, value='円×' if p1_str else '').alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    ws.cell(row=start_r, column=c_m2_start, value='回＝' if c1_str else '').alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    ws.cell(row=start_r, column=c_m3_start, value='円' if t1_str else '').alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    
    # 清掃: c_m3_start 以降のセル内容をクリアして円記号重複を完全防止
    for c_i in range(c_m3_start + 1, c_m3_end + 1):
        ws.cell(row=start_r, column=c_i, value=None)


def format_proxy_address(raw_str):
    """代理人住所の改行時、「住所　」の幅（全角3文字分）をインデントして郵便番号の真下に会社名を配置"""
    if not raw_str:
        return "住所　"
    parts = re.split(r'[\r\n]+|[\s　]{2,}', str(raw_str).strip())
    if len(parts) >= 2:
        return f"住所　{parts[0]}\n　　　{parts[1]}"
    else:
        m = re.match(r'^(.*?(?:[0-9０-９-]+)?)\s+(株式会社.*|有限会社.*|合同会社.*)$', str(raw_str).strip())
        if m:
            return f"住所　{m.group(1)}\n　　　{m.group(2)}"
        return f"住所　{raw_str}"


def get_image_anchors(ws):
    """シート内の画像オブジェクトの配置位置（行・列）を取得"""
    image_coords = set()
    if hasattr(ws, '_images'):
        for img in ws._images:
            anchor = getattr(img, 'anchor', None)
            if hasattr(anchor, '_from'):
                c_idx = anchor._from.col + 1
                r_idx = anchor._from.row + 1
                image_coords.add((r_idx, c_idx))
            elif isinstance(anchor, str):
                image_coords.add(anchor.upper())
    return image_coords


def extract_header_box_digits(ws, row_idx):
    """spotlogの指定行（12:公費負担者, 16:公費受給者, 20:区市町村, 24:受給者）から数字列を抽出"""
    digits = []
    for c in range(14, 40):
        v = ws.cell(row=row_idx, column=c).value
        if v is not None and str(v).strip().isdigit():
            digits.append(str(v).strip())
    return digits


def fill_header_boxes(target_ws, target_row, digits):
    """基準様式の8マスヘッダー欄へ右詰めで数値を入力"""
    box_cols = ['AF', 'AL', 'AR', 'AX', 'BD', 'BJ', 'BP', 'BV']
    for col in box_cols:
        target_ws[f"{col}{target_row}"] = None
    if digits:
        digs = digits[-len(box_cols):] if len(digits) > len(box_cols) else digits
        start_idx = len(box_cols) - len(digs)
        for i, d in enumerate(digs):
            cell = target_ws[f"{box_cols[start_idx + i]}{target_row}"]
            cell.value = str(d)
            cell.alignment = Alignment(horizontal="center", vertical="center")


def detect_special_marks(ws, img_coords):
    """特記事項・保険種別・給付割合の丸囲みを動的判定（丸囲みがない場合は固定丸を付与しない）"""
    res = {
        'shakoku': '1 社国',
        'koukou': '3 後高',
        'kouhi': '2 公費',
        'hon_gai': '2 本外',
        'roku_gai': '4 六外',
        'ie_gai': '6 家外',
        'kou_gai_1': '8 高外一',
        'kou_gai_7': '0 高外7',
        'rate_8': '8',
        'rate_9': '9',
        'rate_10': '10'
    }
    
    for r in range(11, 21):
        for c in range(40, ws.max_column + 1):
            v = str(ws.cell(row=r, column=c).value or '')
            if any(m in v for m in ['○', '●', '◎', '①', '②', '③', '④', '⑤', '⑥', '⑦', '⑧', '⑨', '⑩']):
                if '社国' in v: res['shakoku'] = '① 社国'
                if '後高' in v: res['koukou'] = '③ 後高'
                if '公費' in v: res['kouhi'] = '② 公費'
                if '本外' in v: res['hon_gai'] = '② 本外'
                if '六外' in v: res['roku_gai'] = '④ 六外'
                if '家外' in v: res['ie_gai'] = '⑥ 家外'
                if '高外一' in v or '高外1' in v: res['kou_gai_1'] = '⑧ 高外一'
                if '高外7' in v or '高外0' in v: res['kou_gai_7'] = '⓪ 高外7'

    for (r, c) in img_coords:
        if 10 <= r <= 19:
            if 48 <= c <= 54:
                if r <= 14: res['shakoku'] = '① 社国'
                else: res['kouhi'] = '② 公費'
            elif 55 <= c <= 59:
                res['koukou'] = '③ 後高'
            elif 60 <= c <= 63:
                if r <= 13: res['hon_gai'] = '② 本外'
                elif r <= 15: res['roku_gai'] = '④ 六外'
                else: res['ie_gai'] = '⑥ 家外'
            elif 64 <= c <= 69:
                if r <= 14: res['kou_gai_1'] = '⑧ 高外一'
                else: res['kou_gai_7'] = '⓪ 高外7'
            elif 70 <= c <= 73:
                res['rate_8'] = '⑧'
            elif 74 <= c <= 76:
                res['rate_9'] = '⑨'
            elif 77 <= c <= 82:
                res['rate_10'] = '⑩'
                
    return res


def extract_copayment_ratio_text(ws, marks):
    """一部負担金の割合（1割・2割・3割）を動的判定"""
    for r in range(145, 170):
        for c in range(1, 30):
            v = str(ws.cell(row=r, column=c).value or '')
            if '2割' in v or '２割' in v:
                return "一部負担金（１ 割 ・ ② 割 ・ ３ 割）"
            elif '3割' in v or '３割' in v:
                return "一部負担金（１ 割 ・ ２ 割 ・ ③ 割）"
            elif '1割' in v or '１割' in v:
                return "一部負担金（① 割 ・ ２ 割 ・ ３ 割）"
                
    if '⑧' in marks.get('rate_8', ''):
        return "一部負担金（１ 割 ・ ② 割 ・ ３ 割）"
    elif '⑩' in marks.get('rate_10', ''):
        return "一部負担金（１ 割 ・ ２ 割 ・ ③ 割）"
    elif '⑨' in marks.get('rate_9', ''):
        return "一部負担金（① 割 ・ ２ 割 ・ ３ 割）"
        
    return "一部負担金（１ 割 ・ ２ 割 ・ ３ 割）"


def detect_work_injury(ws, img_coords):
    """業務上・外、第三者行為の有無を動的判定"""
    r1 = any((r, c) in img_coords for r in range(38, 43) for c in range(40, 49))
    r2 = any((r, c) in img_coords for r in range(38, 43) for c in range(50, 59))
    r3 = any((r, c) in img_coords for r in range(38, 43) for c in range(60, 72))
    
    other_val = ""
    for r in range(35, 48):
        for c in range(30, ws.max_column + 1):
            v = str(ws.cell(r, c).value or '')
            if 'その他' in v:
                other_val = v
                break
        if other_val: break
    if not other_val:
        other_val = str(ws["BL40"].value or "")
        
    m = re.search(r'[（(](.*?)[)）]', other_val)
    inside_txt = m.group(1) if m else "　不　詳　"
    if not inside_txt.strip():
        inside_txt = "　不　詳　"
        
    if r1:
        return f"（　①．業務上　２．第三者行為　３．その他（　　　　））"
    elif r2:
        return f"（　１．業務上　②．第三者行為　３．その他（　　　　））"
    elif r3 or "その他" in other_val:
        return f"（　１．業務上　２．第三者行為　③．その他（{inside_txt}））"
    else:
        return f"（　１．業務上　２．第三者行為　３．その他（{inside_txt}））"


def detect_claim_type(ws, img_coords):
    """請求区分（新規・継続）を動的判定"""
    is_new = any((r, c) in img_coords for r in range(52, 59) for c in range(60, 69))
    is_cont = any((r, c) in img_coords for r in range(52, 59) for c in range(70, 78))
    if is_new:
        return "○新規・継続"
    elif is_cont:
        return "新規・○継続"
    return "新規・○継続"


def detect_outcome(ws, img_coords):
    """転帰（継続・治癒・中止・転医）を動的判定"""
    is_cont = any((r, c) in img_coords for r in range(60, 68) for c in range(60, 66))
    is_cure = any((r, c) in img_coords for r in range(60, 68) for c in range(67, 72))
    is_stop = any((r, c) in img_coords for r in range(60, 68) for c in range(73, 78))
    is_trans = any((r, c) in img_coords for r in range(60, 68) for c in range(79, 85))
    if is_cure:
        return "継続・○治癒・中止・転医"
    elif is_stop:
        return "継続・治癒・○中止・転医"
    elif is_trans:
        return "継続・治癒・中止・○転医"
    else:
        return "○継続・治癒・中止・転医"


def detect_visit_reasons(ws, img_coords):
    """往療又は訪問の理由を動的判定（複数選択 1と2両方、その他 に完全対応）"""
    target_row = None
    for r in range(160, 185):
        for c in range(1, 15):
            v = str(ws.cell(row=r, column=c).value or '')
            if "往療又は訪問の理由" in v:
                target_row = r
                break
        if target_row:
            break
            
    if not target_row:
        target_row = 167

    # 画像の丸囲み判定
    r1 = any((r, c) in img_coords for r in range(target_row - 1, target_row + 2) for c in range(15, 25))
    r2 = any((r, c) in img_coords for r in range(target_row - 1, target_row + 2) for c in range(34, 44))
    r3 = any((r, c) in img_coords for r in range(target_row - 1, target_row + 2) for c in range(60, 70))

    # テキスト内の○/①/②/③判定
    for c in range(10, ws.max_column + 1):
        v = str(ws.cell(row=target_row, column=c).value or '')
        if '○１' in v or '①' in v or '○1' in v: r1 = True
        if '○２' in v or '②' in v or '○2' in v: r2 = True
        if '○３' in v or '③' in v or '○3' in v: r3 = True

    other_txt = "　　　"
    bm_val = str(ws.cell(row=target_row, column=65).value or '')
    m = re.search(r'[（(](.*?)[)）]', bm_val)
    if m and m.group(1).strip():
        other_txt = m.group(1)

    opt1_str = "①．独歩による公共交通機関を使っての外出困難" if r1 else "１．独歩による公共交通機関を使っての外出困難"
    opt2_str = "②．認知症や視覚、内部、精神障害などにより独歩による外出困難" if r2 else "２．認知症や視覚、内部、精神障害などにより独歩による外出困難"
    opt3_str = f"③．その他（{other_txt}）" if r3 else f"３．その他（{other_txt}）"

    return f"○往療又は訪問の理由（ {opt1_str}　{opt2_str}　{opt3_str} ）"


def detect_practitioner_location_type(ws, img_coords):
    """施術証明欄の「1.施術所所在地」「2.出張専門施術者住所地」を動的判定"""
    is_house_call = any((r, c) in img_coords for r in range(168, 186) for c in range(60, 75))
    if is_house_call:
        return "1.施術所所在地　②.出張専門施術者住所地"
    return "①.施術所所在地　2.出張専門施術者住所地"


def detect_electrotherapy(ws, img_coords):
    """電療料（加算／ 1電気針 2電気温灸器 3電気光線器具）の丸囲みを動的判定"""
    r_found = None
    for r in range(110, 145):
        for c in range(1, 20):
            v = str(ws.cell(row=r, column=c).value or '').replace(' ', '')
            if '電療料' in v:
                r_found = r
                break
        if r_found: break
    if not r_found: r_found = 125
    
    opt1 = any((r_i, c_i) in img_coords for r_i in range(r_found - 2, r_found + 3) for c_i in range(15, 23))
    opt2 = any((r_i, c_i) in img_coords for r_i in range(r_found - 2, r_found + 3) for c_i in range(23, 30))
    opt3 = any((r_i, c_i) in img_coords for r_i in range(r_found - 2, r_found + 3) for c_i in range(30, 39))
    
    for c in range(1, ws.max_column + 1):
        val = str(ws.cell(row=r_found, column=c).value or '')
        if '○１' in val or '①' in val or '○1' in val: opt1 = True
        if '○２' in val or '②' in val or '○2' in val: opt2 = True
        if '○３' in val or '③' in val or '○3' in val: opt3 = True
    
    s1 = '① 電気針' if opt1 else '１電気針'
    s2 = '② 電気温灸器' if opt2 else '２電気温灸器'
    s3 = '③ 電気光線器具' if opt3 else '３電気光線器具'
    return f"電療料（加算／　{s1}　{s2}　{s3}）"


def extract_report_prev_date(ws, is_massage=False):
    """施術報告書交付料（前回支給：　年　月分）の年月テキストを動的抽出"""
    target_row = None
    for r in range(120, 175):
        for c in range(1, 20):
            v = str(ws.cell(row=r, column=c).value or '').replace(' ', '')
            if '施術報告書' in v or '報告書交付' in v:
                target_row = r
                break
        if target_row: break
        
    if not target_row:
        target_row = 152 if is_massage else 137
        
    for c in range(10, 40):
        v = ws.cell(row=target_row, column=c).value
        if v and str(v).strip():
            str_v = str(v).strip()
            if re.search(r'\d+', str_v) and not str_v.isdigit():
                clean_str = str_v.replace('前回支給:', '').replace('(', '').replace(')', '').strip()
                if clean_str:
                    return clean_str
            elif re.search(r'\d+\s*年\s*\d+\s*月', str_v):
                return str_v.strip()
    return "　年　月分"


def check_is_ineligible_for_standard_form(ws, sheet_type):
    """
    基準様式への変換対象外（月16回以降の50%逓減、訪問施術料4・5、集中率80%逓減など）を厳密判定
    複数該当する場合はすべての理由を結合して返却
    返り値: (is_ineligible: bool, reason: str)
    """
    reasons = []
    
    # 1. 訪問施術料4 (10〜19人) / 5 (20人以上) の判定
    has_fee_4_5 = False
    fee_4_rows = [109, 117] if sheet_type == '鍼灸' else [106, 114]
    for r in fee_4_rows:
        for c in range(20, 60):
            v = ws.cell(row=r, column=c).value
            if v and str(v).strip() not in ['', '0', '0回', '0円']:
                try:
                    num = float(str(v).replace(',', '').replace('円', '').replace('回', '').strip())
                    if num > 0:
                        has_fee_4_5 = True
                        break
                except Exception:
                    pass
        if has_fee_4_5:
            break
    if has_fee_4_5:
        reasons.append("訪問施術料4または5（10人以上）")

    # 2. 月16回以降の50%逓減の算定有無判定 (通所、訪問1〜3、加算等すべて)
    has_50_teigen = False
    for r in range(70, 165):
        is_teigen_row = False
        for c in range(1, 20):
            cell_v = str(ws.cell(row=r, column=c).value or '')
            if any(k in cell_v for k in ['50％逓減', '50%逓減', '月16回以降', '16回以降']):
                is_teigen_row = True
                break
        
        if is_teigen_row:
            for c in range(20, 65):
                v = ws.cell(row=r, column=c).value
                if v and str(v).strip() not in ['', '0', '0回', '0円']:
                    try:
                        num = float(str(v).replace(',', '').replace('円', '').replace('回', '').strip())
                        if num > 0:
                            has_50_teigen = True
                            break
                    except Exception:
                        pass
        if has_50_teigen:
            break
    if has_50_teigen:
        reasons.append("月16回以降の施術（50%逓減）")

    # 3. 集中率90%以上（80%逓減対象）の判定
    has_80_teigen = False
    for r in range(135, 170):
        is_80_row = False
        for c in range(1, 25):
            v = str(ws.cell(row=r, column=c).value or '')
            if any(k in v for k in ['80％逓減', '80%逓減', '集中率90', '集中率９０']):
                is_80_row = True
                break
        if is_80_row:
            for c in range(20, ws.max_column + 1):
                val = ws.cell(row=r, column=c).value
                if val and str(val).strip() not in ['', '0', '0円', '逓減後合計額']:
                    try:
                        num = float(str(val).replace(',', '').replace('円', '').strip())
                        if num > 0:
                            has_80_teigen = True
                            break
                    except Exception:
                        pass
        if has_80_teigen:
            break
    if has_80_teigen:
        reasons.append("施設集中率90％以上（80%逓減）")

    if reasons:
        reason_str = "、".join(reasons) + "が算定されているため"
        return True, reason_str
    return False, ""


def detect_payment_section(ws, img_coords):
    """支払機関欄（支払区分・預金の種類・金融機関種別・支店種別・名称）を動的判定"""
    r_start = None
    for r in range(195, 230):
        for c in range(1, 10):
            v = str(ws.cell(row=r, column=c).value or '')
            if '支払機関' in v or '支払区分' in v:
                r_start = r
                break
        if r_start:
            break
            
    if not r_start:
        r_start = 203

    bank_name = str(ws.cell(row=r_start, column=53).value or '').strip()
    bank_type_raw = str(ws.cell(row=r_start, column=63).value or '').strip()
    branch_name = str(ws.cell(row=r_start, column=67).value or '').strip()
    branch_type_raw = str(ws.cell(row=r_start, column=75).value or '').strip()
    
    pay_1 = any((r, c) in img_coords for r in range(r_start+1, r_start+5) for c in range(5, 15))
    pay_2 = any((r, c) in img_coords for r in range(r_start+1, r_start+5) for c in range(16, 26))
    pay_3 = any((r, c) in img_coords for r in range(r_start+4, r_start+8) for c in range(5, 15))
    pay_4 = any((r, c) in img_coords for r in range(r_start+4, r_start+8) for c in range(16, 26))
    if not any([pay_1, pay_2, pay_3, pay_4]):
        pay_1 = True
        
    dep_1 = any((r, c) in img_coords for r in range(r_start+1, r_start+5) for c in range(27, 36))
    dep_2 = any((r, c) in img_coords for r in range(r_start+1, r_start+5) for c in range(36, 46))
    dep_3 = any((r, c) in img_coords for r in range(r_start+4, r_start+8) for c in range(27, 36))
    dep_4 = any((r, c) in img_coords for r in range(r_start+4, r_start+8) for c in range(36, 46))
    if not any([dep_1, dep_2, dep_3, dep_4]):
        dep_1 = True

    return {
        'bank_name': bank_name,
        'bank_type': bank_type_raw or '銀行',
        'branch_name': branch_name,
        'branch_type': branch_type_raw or '支店',
        'pay_1': pay_1,
        'pay_2': pay_2,
        'pay_3': pay_3,
        'pay_4': pay_4,
        'dep_1': dep_1,
        'dep_2': dep_2,
        'dep_3': dep_3,
        'dep_4': dep_4
    }


def detect_acupuncture_diseases(ws, img_coords):
    """はり・きゅうの傷病名を動的判定（複数選択に完全対応・①.形式で配置）"""
    diseases = {
        'd1': '1. 神経痛',
        'd2': '2. リウマチ',
        'd3': '3. 頸腕症候群',
        'd4': '4. 五十肩',
        'd5': '5. 腰痛症',
        'd6': '6. 頸椎捻挫後遺症',
        'd7': '7. その他（　　　　　　　　　）'
    }
    
    for (r, c) in img_coords:
        if 58 <= r <= 68:
            if 14 <= c <= 22:
                if r <= 63: diseases['d1'] = '①. 神経痛'
                else: diseases['d5'] = '⑤. 腰痛症'
            elif 23 <= c <= 35:
                if r <= 63: diseases['d2'] = '②. リウマチ'
                else: diseases['d6'] = '⑥. 頸椎捻挫後遺症'
            elif 36 <= c <= 46:
                if r <= 63: diseases['d3'] = '③. 頸腕症候群'
                else: diseases['d7'] = '⑦. その他（　　　　　　　　　）'
            elif 47 <= c <= 58:
                if r <= 63: diseases['d4'] = '④. 五十肩'

    for r in range(55, 70):
        for c in range(10, ws.max_column + 1):
            v = str(ws.cell(row=r, column=c).value or '')
            if any(m in v for m in ['○', '●', '◎', '①', '②', '③', '④', '⑤', '⑥', '⑦']):
                if '神経痛' in v: diseases['d1'] = '①. 神経痛'
                if 'リウマチ' in v: diseases['d2'] = '②. リウマチ'
                if '頸腕' in v: diseases['d3'] = '③. 頸腕症候群'
                if '五十肩' in v: diseases['d4'] = '④. 五十肩'
                if '腰痛' in v: diseases['d5'] = '⑤. 腰痛症'
                if '頸椎' in v: diseases['d6'] = '⑥. 頸椎捻挫後遺症'
                if 'その他' in v: diseases['d7'] = '⑦. その他（　　　　　　　　　）'

    # 上部の傷病名欄テキストからも補助検知
    for r in range(25, 40):
        for c in range(20, min(ws.max_column + 1, 60)):
            v = str(ws.cell(row=r, column=c).value or '')
            if any(k in v for k in ['神経痛', '五十肩', '腰痛', 'リウマチ', '頸腕', '頸椎']):
                if '神経痛' in v: diseases['d1'] = '①. 神経痛'
                if 'リウマチ' in v: diseases['d2'] = '②. リウマチ'
                if '頸腕' in v: diseases['d3'] = '③. 頸腕症候群'
                if '五十肩' in v: diseases['d4'] = '④. 五十肩'
                if '腰痛' in v: diseases['d5'] = '⑤. 腰痛症'
                if '頸椎' in v: diseases['d6'] = '⑥. 頸椎捻挫後遺症'
        
    return diseases


def detect_first_exam_fee(spot_ws, img_coords):
    """初検料（１はり ２きゅう ３はりきゅう併用）の選択肢および金額を動的判定"""
    r_found = None
    for r in range(60, 80):
        for c in range(1, 20):
            v = str(spot_ws.cell(row=r, column=c).value or '')
            if '初検' in v:
                r_found = r
                break
        if r_found:
            break
            
    if not r_found:
        return {'text': '初検料（１はり　２きゅう　３はりきゅう併用）', 'price': ''}
        
    price_val = None
    for c in range(25, 60):
        v = spot_ws.cell(row=r_found, column=c).value
        if v is not None and str(v).strip().isdigit() and int(str(v).strip()) > 0:
            price_val = int(str(v).strip())
            break
            
    opt1, opt2, opt3 = False, False, False
    for c in range(1, spot_ws.max_column + 1):
        v = str(spot_ws.cell(row=r_found, column=c).value or '')
        if '①' in v or '○1' in v or '○１' in v or '1はり' in v: opt1 = True
        if '②' in v or '○2' in v or '○２' in v or '2きゅう' in v: opt2 = True
        if '③' in v or '○3' in v or '○３' in v or '併用' in v: opt3 = True
        
    s1 = '①はり' if opt1 else '１はり'
    s2 = '②きゅう' if opt2 else '２きゅう'
    s3 = '③はりきゅう併用' if opt3 else '３はりきゅう併用'
    
    text = f'初検料（{s1}　{s2}　{s3}）'
    return {'text': text, 'price': f'{price_val:,}' if price_val else ''}


def extract_header_data_dynamic(spot_ws):
    """保険者番号・被保険者記号番号・発病年月日・原因経過・氏名・フリガナ・性別・続柄を行ズレに関わらず動的抽出"""
    # 1. 保険者番号 (8桁マス目用数字リスト)
    ins_digits = []
    for r in range(15, 32):
        for c in range(20, spot_ws.max_column + 1):
            v = str(spot_ws.cell(row=r, column=c).value or '')
            if '保険者番号' in v.replace(' ', ''):
                for col_k in range(c + 1, min(spot_ws.max_column + 1, c + 45)):
                    cv = spot_ws.cell(row=r, column=col_k).value
                    if cv is not None:
                        digits = re.findall(r'\d', str(cv))
                        ins_digits.extend(digits)
                break
        if ins_digits:
            break

    # 2. 記号番号, 発病負傷, 原因経過
    kigou, hatsubyou, genin = None, None, None
    for r in range(25, 38):
        for c in range(1, spot_ws.max_column + 1):
            v = str(spot_ws.cell(row=r, column=c).value or '')
            clean_v = v.replace(' ', '').replace('\n', '')
            if '被保険者' in clean_v and ('記号' in clean_v or '番号' in clean_v):
                for ro in range(r + 1, r + 5):
                    cv = spot_ws.cell(row=ro, column=c).value or spot_ws.cell(row=ro, column=c + 3).value
                    if cv:
                        kigou = str(cv).strip()
                        break
            if '発病' in clean_v or '負傷年月日' in clean_v:
                for ro in range(r + 1, r + 5):
                    cv = spot_ws.cell(row=ro, column=c).value or spot_ws.cell(row=ro, column=c + 3).value
                    if cv:
                        hatsubyou = str(cv).strip()
                        break
            if '原因及びその経過' in clean_v or '原因及び経過' in clean_v or '発症又は負傷' in clean_v:
                for ro in range(r + 1, r + 5):
                    cv = spot_ws.cell(row=ro, column=c).value or spot_ws.cell(row=ro, column=c + 3).value
                    if cv:
                        genin = str(cv).strip()
                        break

    # 3. 氏名, フリガナ, 性別, 続柄
    kana, name, gender, relation = None, None, None, None
    for r in range(32, 48):
        for c in range(1, spot_ws.max_column + 1):
            v = str(spot_ws.cell(row=r, column=c).value or '')
            if '(ﾌﾘｶﾞﾅ)' in v or 'フリガナ' in v:
                for co in range(c + 1, c + 15):
                    cv = spot_ws.cell(row=r, column=co).value
                    if cv and str(cv).strip():
                        kana = str(cv).strip()
                        break
                for ro in range(r + 1, r + 5):
                    for co in range(c - 2, c + 15):
                        cv = spot_ws.cell(row=ro, column=co).value
                        if cv and len(str(cv).strip()) >= 2 and not any(k in str(cv) for k in ['男', '女', '本人', '家族', '続柄']):
                            name = str(cv).strip()
                            break
                    if name:
                        break
            if '続柄' in v.replace(' ', ''):
                for ro in range(r + 1, r + 5):
                    for co in range(c - 2, c + 5):
                        cv = spot_ws.cell(row=ro, column=co).value
                        if cv and any(k in str(cv) for k in ['本人', '家族', '妻', '夫', '子', '父', '母']):
                            relation = str(cv).strip()
                            break
                    if relation:
                        break
            if v in ['男', '女']:
                gender = v

    return {
        'ins_digits': ins_digits,
        'kigou': kigou,
        'hatsubyou': hatsubyou,
        'genin': genin,
        'kana': kana,
        'name': name,
        'gender': gender,
        'relation': relation
    }


def extract_calendar_marks_dynamic(spot_ws):
    """カレンダーの「1〜31」日付行と各日の丸印（①〜⑤、○、●、◎）を行ズレに関わらず動的抽出"""
    cal_data = {}
    day_row = None
    col_day_map = {}
    
    # 1. 「1〜31」の日付が並ぶ行を自動探索
    for r in range(140, 200):
        found_days = {}
        for c in range(10, min(spot_ws.max_column + 1, 90)):
            v = spot_ws.cell(row=r, column=c).value
            if v is not None and str(v).strip().isdigit():
                d = int(str(v).strip())
                if 1 <= d <= 31:
                    found_days[c] = d
        if len(found_days) >= 10:
            day_row = r
            col_day_map = found_days
            break
            
    # 2. 日付行の下の行をスキャンして丸印を抽出
    if day_row:
        for r in range(day_row + 1, min(spot_ws.max_row + 1, day_row + 10)):
            for c, d in col_day_map.items():
                if d not in cal_data:
                    v = spot_ws.cell(row=r, column=c).value
                    if v and str(v).strip() in ['①', '②', '③', '④', '⑤', '◎', '○', '●', '1', '2', '3', '4', '5']:
                        mark_str = str(v).strip()
                        num_to_enc = {'1': '①', '2': '②', '3': '③', '4': '④', '5': '⑤'}
                        cal_data[d] = num_to_enc.get(mark_str, mark_str)
                        
    return cal_data


def clean_amount(val):
    if val is None: return None
    if isinstance(val, (int, float)):
        if val == 0: return None
        return int(round(val))
    s = str(val).strip()
    if s in ['', '円', '0', '0円', '0.0']: return None
    s_clean = s.replace('円', '').replace(',', '').strip()
    try:
        f_val = float(s_clean)
        if f_val == 0: return None
        return int(round(f_val))
    except ValueError:
        pass
    digits = re.sub(r'[^\d]', '', s)
    if digits: return int(digits)
    return None


def parse_fee_row_smart(spot_ws, r):
    """行内の『円×』『回＝』『円』マーカーの位置を基準に、単価・回数・金額をタテ・ヨコ完全動的探索"""
    row_cells = [(c, spot_ws.cell(r, c).value) for c in range(1, spot_ws.max_column + 1) if spot_ws.cell(r, c).value is not None]
    price = None
    count = None
    total = None
    col_yen_x = None
    col_kai_eq = None
    col_yen_end = None
    
    for c, v in row_cells:
        s = str(v).replace(' ', '').replace('　', '')
        if '円×' in s or '円*' in s:
            col_yen_x = c
        elif '回＝' in s or '回=' in s:
            col_kai_eq = c
        elif s == '円' and col_kai_eq and c > col_kai_eq:
            col_yen_end = c
            
    if col_yen_x:
        for c in range(col_yen_x - 1, 20, -1):
            val = clean_amount(spot_ws.cell(r, c).value)
            if val is not None:
                price = val
                break
                
    if col_kai_eq:
        start_c = col_yen_x if col_yen_x else 20
        for c in range(col_kai_eq - 1, start_c, -1):
            val = clean_amount(spot_ws.cell(r, c).value)
            if val is not None:
                count = val
                break
                
    if col_kai_eq:
        end_c = col_yen_end if col_yen_end else spot_ws.max_column + 1
        for c in range(col_kai_eq + 1, end_c):
            val = clean_amount(spot_ws.cell(r, c).value)
            if val is not None:
                total = val
                break
    else:
        for c in range(25, spot_ws.max_column + 1):
            val = clean_amount(spot_ws.cell(r, c).value)
            if val is not None:
                total = val
                break
                
    if total is None and count is None:
        price = None
        
    return {'price': price, 'count': count, 'total': total}


def extract_middle_fees_dynamic(spot_ws):
    """通所、訪問施術料1〜3、電療料、温罨法、特別地域、往療料、施術報告書交付料、明細書発行加算、合計、一部負担金、請求額をタテ・ヨコ完全動的抽出"""
    fees = {}
    for r in range(65, 175):
        for c in range(1, 20):
            v = str(spot_ws.cell(row=r, column=c).value or '').replace(' ', '').replace('　', '')

            if '通所' in v and 'tuusho' not in fees:
                u_row = parse_fee_row_smart(spot_ws, r)
                l_row = parse_fee_row_smart(spot_ws, r+2)
                fees['tuusho'] = {
                    'u_price': u_row['price'],
                    'u_count': u_row['count'],
                    'u_total': u_row['total'],
                    'l_price': l_row['price'],
                    'l_count': l_row['count'],
                    'l_total': l_row['total'],
                }
            elif ('訪問施術料１' in v or '訪問施術料1' in v) and 'h1' not in fees:
                u_row = parse_fee_row_smart(spot_ws, r)
                l_row = parse_fee_row_smart(spot_ws, r+2)
                fees['h1'] = {
                    'u_price': u_row['price'],
                    'u_count': u_row['count'],
                    'u_total': u_row['total'],
                    'l_price': l_row['price'],
                    'l_count': l_row['count'],
                    'l_total': l_row['total'],
                }
            elif ('訪問施術料２' in v or '訪問施術料2' in v) and 'h2' not in fees:
                u_row = parse_fee_row_smart(spot_ws, r)
                l_row = parse_fee_row_smart(spot_ws, r+2)
                fees['h2'] = {
                    'u_price': u_row['price'],
                    'u_count': u_row['count'],
                    'u_total': u_row['total'],
                    'l_price': l_row['price'],
                    'l_count': l_row['count'],
                    'l_total': l_row['total'],
                }
            elif ('訪問施術料３' in v or '訪問施術料3' in v) and 'h3' not in fees:
                u_row = parse_fee_row_smart(spot_ws, r)
                l_row = parse_fee_row_smart(spot_ws, r+2)
                fees['h3'] = {
                    'u_price': u_row['price'],
                    'u_count': u_row['count'],
                    'u_total': u_row['total'],
                    'l_price': l_row['price'],
                    'l_count': l_row['count'],
                    'l_total': l_row['total'],
                }
            elif ('温罨法・電気光線' in v or '温罨法電気' in v) and 'on_den' not in fees:
                fees['on_den'] = parse_fee_row_smart(spot_ws, r)
            elif '温罨法' in v and 'on_an' not in fees and '電気' not in v:
                fees['on_an'] = parse_fee_row_smart(spot_ws, r)
            elif '電療料' in v and 'denryou' not in fees:
                fees['denryou'] = parse_fee_row_smart(spot_ws, r)
            elif '特別地域' in v and 'tokubetsu' not in fees:
                fees['tokubetsu'] = parse_fee_row_smart(spot_ws, r)
            elif '往療料' in v and 'ouryou' not in fees:
                fees['ouryou'] = parse_fee_row_smart(spot_ws, r)
            elif ('施術報告書' in v or '報告書交付' in v) and 'houkoku' not in fees:
                fees['houkoku'] = parse_fee_row_smart(spot_ws, r)
            elif ('明細書' in v or '明細書発行' in v) and 'meisai' not in fees:
                fees['meisai'] = parse_fee_row_smart(spot_ws, r)
            elif ('合計' in v or '合　計' in v) and 'goukei' not in fees:
                parsed = parse_fee_row_smart(spot_ws, r)
                fees['goukei'] = parsed['total']
            elif '一部負担金' in v and 'ichibu' not in fees:
                parsed = parse_fee_row_smart(spot_ws, r)
                fees['ichibu'] = parsed['total']
            elif ('請求額' in v or '請　求額' in v) and 'seikyuu' not in fees:
                parsed = parse_fee_row_smart(spot_ws, r)
                fees['seikyuu'] = parsed['total']

    return fees


def extract_cert_section_dynamic(spot_ws):
    """施術証明欄の郵便番号、年月日、所在地、施術所名、登録記号番号、管理者名、電話番号を動的抽出"""
    r_cert = None
    for r in range(140, 220):
        for c in range(1, 15):
            v = str(spot_ws.cell(row=r, column=c).value or '').replace(' ', '').replace('　', '')
            if '施術証明' in v or '上記のとおり施術を行い' in v:
                r_cert = r
                break
        if r_cert: break
    if not r_cert: return {}
    
    zip_val, date_val, addr_val, clinic_name, reg_no, manager_name, tel_val = None, None, None, None, None, None, None
    
    for r in range(r_cert, min(r_cert + 25, spot_ws.max_row + 1)):
        for c in range(1, spot_ws.max_column + 1):
            raw_v = spot_ws.cell(row=r, column=c).value
            if raw_v is None: continue
            raw_str = str(raw_v).strip()
            v = raw_str.replace(' ', '').replace('　', '')
            
            if ('〒' in raw_str or re.search(r'\b\d{3}-\d{4}\b', raw_str)) and not zip_val:
                zip_val = raw_str
                
            if ('令和' in raw_str or '平成' in raw_str) and '年' in raw_str and '月' in raw_str and '日' in raw_str and c < 30 and not date_val:
                date_val = raw_str
                
            if v == '所在地':
                for co in range(c + 1, spot_ws.max_column + 1):
                    av = spot_ws.cell(row=r, column=co).value
                    if av and str(av).strip() and str(av).strip() not in ['所在地', '住所', '施術所', ':']:
                        addr_val = str(av).strip()
                        break
                        
            if ('登録記号番号' in v or '登録番号' in v) and not reg_no:
                for r_down in range(r + 1, r + 6):
                    for c_near in range(max(1, c - 2), c + 6):
                        rv = spot_ws.cell(row=r_down, column=c_near).value
                        if rv and str(rv).strip():
                            s_rv = str(rv).strip()
                            if not any(k in s_rv for k in ['登録', '名称', '氏名', '管理者', '電話', '〒', '所在地', '住所']):
                                reg_no = s_rv
                                break
                    if reg_no: break
                    
            if v in ['名称', '名・称', '施術所名', '施術所の名称'] and not clinic_name:
                for co in range(c + 1, spot_ws.max_column + 1):
                    cv = spot_ws.cell(row=r, column=co).value
                    if cv and str(cv).strip() and str(cv).strip() not in ['名称', ':']:
                        clinic_name = str(cv).strip()
                        break
                        
            if ('施術管理者' in v or '管理者' in v) and not manager_name:
                for co in range(c + 1, spot_ws.max_column + 1):
                    mv = spot_ws.cell(row=r, column=co).value
                    if mv and str(mv).strip() and str(mv).strip() not in ['氏名', '氏　名', '管理者', '施術管理者', '電話', ':']:
                        manager_name = str(mv).strip()
                        break
                        
            if ('電話' in v or 'TEL' in v.upper()) and not tel_val:
                for co in range(c + 1, spot_ws.max_column + 1):
                    tv = spot_ws.cell(row=r, column=co).value
                    if tv and str(tv).strip() and re.search(r'[\d-]{8,}', str(tv).strip()):
                        tel_val = str(tv).strip()
                        break
                        
    return {
        'zip': zip_val, 'date': date_val, 'addr': addr_val,
        'clinic_name': clinic_name, 'reg_no': reg_no,
        'manager_name': manager_name, 'tel': tel_val
    }


def extract_applicant_section_dynamic(spot_ws):
    """申請欄の郵便番号、申請日、住所、広域連合名、申請者氏名、電話番号を動的抽出"""
    r_app = None
    for r in range(175, 225):
        for c in range(1, 15):
            v = str(spot_ws.cell(row=r, column=c).value or '').replace(' ', '').replace('　', '')
            if '申請欄' in v or '上記の療養に要した費用' in v:
                r_app = r
                break
        if r_app: break
    if not r_app: return {}
    
    zip_val, date_val, addr_val, kouiki_val, name_val, tel_val = None, None, None, None, None, None
    
    for r in range(r_app, min(r_app + 20, spot_ws.max_row + 1)):
        for c in range(1, spot_ws.max_column + 1):
            raw_v = spot_ws.cell(row=r, column=c).value
            if raw_v is None: continue
            raw_str = str(raw_v).strip()
            v = raw_str.replace(' ', '').replace('　', '')
            
            if ('〒' in raw_str or re.search(r'\b\d{3}-\d{4}\b', raw_str)) and not zip_val:
                zip_val = raw_str
                
            if ('令和' in raw_str or '平成' in raw_str) and '年' in raw_str and '月' in raw_str and '日' in raw_str and c < 30 and not date_val:
                date_val = raw_str
                
            if '殿' in raw_str or '広域連合' in raw_str or '健康保険組合' in raw_str:
                kouiki_val = raw_str
                
            if v == '住所' and not addr_val:
                for co in range(c + 1, spot_ws.max_column + 1):
                    av = spot_ws.cell(row=r, column=co).value
                    if av and str(av).strip() and str(av).strip() not in ['住所', '申請者', '（被保険者）', ':']:
                        addr_val = str(av).strip()
                        break
                        
            if v in ['氏名', '氏・名'] and not name_val:
                for co in range(c + 1, spot_ws.max_column + 1):
                    nv = spot_ws.cell(row=r, column=co).value
                    if nv and str(nv).strip() and str(nv).strip() not in ['氏名', '氏　名', '住所', '申請者', '（被保険者）', '電話', ':']:
                        name_val = str(nv).strip()
                        break
                        
            if ('電話' in v or 'TEL' in v.upper()) and not tel_val:
                for co in range(c + 1, spot_ws.max_column + 1):
                    tv = spot_ws.cell(row=r, column=co).value
                    if tv and str(tv).strip() and re.search(r'[\d-]{8,}', str(tv).strip()):
                        tel_val = str(tv).strip()
                        break
                        
    return {
        'zip': zip_val, 'date': date_val, 'addr': addr_val,
        'kouiki': kouiki_val, 'name': name_val, 'tel': tel_val
    }


def extract_delegation_section_dynamic(spot_ws):
    """代理人受領委任欄の委任日、申請者住所・氏名、代理人（足森様）住所・氏名を動的抽出"""
    r_del = None
    for r in range(210, 255):
        for c in range(1, 15):
            v = str(spot_ws.cell(row=r, column=c).value or '').replace(' ', '').replace('　', '')
            if '本申請書に基づく給付金' in v or '受領を代理人に委任' in v:
                r_del = r
                break
        if r_del: break
    if not r_del: return {}
    
    date_val = None
    applicant_addr = None
    delegate_addr = None
    applicant_name = None
    delegate_name = None
    
    for c in range(35, spot_ws.max_column + 1):
        dv = spot_ws.cell(row=r_del, column=c).value
        if dv and ('令和' in str(dv) or '平成' in str(dv)):
            date_val = str(dv).strip()
            break
            
    for r in range(r_del + 1, min(r_del + 20, spot_ws.max_row + 1)):
        for c in range(1, spot_ws.max_column + 1):
            raw_v = spot_ws.cell(row=r, column=c).value
            if raw_v is None: continue
            raw_str = str(raw_v).strip()
            v = raw_str.replace(' ', '').replace('　', '')
            
            if v == '住所':
                for co in range(c + 1, spot_ws.max_column + 1):
                    av = spot_ws.cell(row=r, column=co).value
                    if av and str(av).strip() and str(av).strip() not in ['住所', '代理人', '申請者']:
                        if c < 40 and not applicant_addr:
                            applicant_addr = str(av).strip()
                        elif c >= 40 and not delegate_addr:
                            delegate_addr = str(av).strip()
                        break
                        
            if v == '氏名':
                for co in range(c + 1, spot_ws.max_column + 1):
                    nv = spot_ws.cell(row=r, column=co).value
                    if nv and str(nv).strip() and str(nv).strip() not in ['氏名', '氏　名', '住所', '代理人', '申請者', '（被保険者）']:
                        if c < 40 and not applicant_name:
                            applicant_name = str(nv).strip()
                        elif c >= 40 and not delegate_name:
                            delegate_name = str(nv).strip()
                        break

    return {
        'date': date_val,
        'applicant_addr': applicant_addr,
        'delegate_addr': delegate_addr,
        'applicant_name': applicant_name,
        'delegate_name': delegate_name
    }


def extract_massage_body_counts_dynamic(spot_ws):
    """マッサージの同意部位別施術回数（躯幹・右上肢・左上肢・右下肢・左下肢）および変形徒手矯正術回数を動的抽出"""
    kukan, r_arm, l_arm, r_leg, l_leg = '', '', '', '', ''
    for r in range(50, 100):
        for c in range(1, 40):
            v = str(spot_ws.cell(r, c).value or '').replace(' ', '').replace('　', '')
            if 'マッサージ（施術料）' in v or 'マッサージ(施術料)' in v or v == '同意部位':
                for r_cnt in range(r, r + 6):
                    row_vals = [(col, str(spot_ws.cell(r_cnt, col).value or '').strip()) for col in range(1, spot_ws.max_column + 1) if spot_ws.cell(r_cnt, col).value is not None]
                    if any('施術回数' in val for col, val in row_vals):
                        counts = [val for col, val in row_vals if '回' in val and '施術回数' not in val and val != '']
                        if len(counts) >= 5:
                            kukan, r_arm, l_arm, r_leg, l_leg = counts[0], counts[1], counts[2], counts[3], counts[4]
                        elif len(counts) > 0:
                            for col, val in row_vals:
                                if '回' in val and '施術回数' not in val:
                                    if col < 40 and not kukan: kukan = val
                                    elif col < 46 and not r_arm: r_arm = val
                                    elif col < 52 and not l_arm: l_arm = val
                                    elif col < 58 and not r_leg: r_leg = val
                                    elif not l_leg: l_leg = val
                        break
                if kukan or r_arm or l_arm or r_leg or l_leg: break
        if kukan or r_arm or l_arm or r_leg or l_leg: break
        
    henkei_r_arm, henkei_l_arm, henkei_r_leg, henkei_l_leg = '', '', '', ''
    for r in range(110, 160):
        for c in range(1, 40):
            v = str(spot_ws.cell(r, c).value or '').replace(' ', '').replace('　', '')
            if '変形徒手' in v:
                for r_cnt in range(r, r + 6):
                    row_vals = [(col, str(spot_ws.cell(r_cnt, col).value or '').strip()) for col in range(1, spot_ws.max_column + 1) if spot_ws.cell(r_cnt, col).value is not None]
                    if any('施術回数' in val for col, val in row_vals):
                        counts = [val for col, val in row_vals if '回' in val and '施術回数' not in val and val != '']
                        if len(counts) >= 4:
                            henkei_r_arm, henkei_l_arm, henkei_r_leg, henkei_l_leg = counts[0], counts[1], counts[2], counts[3]
                        break
                if henkei_r_arm or henkei_l_arm: break
        if henkei_r_arm or henkei_l_arm: break
        
    return {
        'kukan': kukan, 'r_arm': r_arm, 'l_arm': l_arm, 'r_leg': r_leg, 'l_leg': l_leg,
        'henkei_r_arm': henkei_r_arm, 'henkei_l_arm': henkei_l_arm, 'henkei_r_leg': henkei_r_leg, 'henkei_l_leg': henkei_l_leg
    }


def extract_acupuncture_jutsu_counts_dynamic(spot_ws):
    """はり・きゅうの1術・2術回数および摘要欄テキストを動的抽出"""
    jutsu1 = None
    jutsu2 = None
    summary_text = None
    
    for r in range(50, 100):
        for c in range(1, 40):
            v = str(spot_ws.cell(r, c).value or '').replace(' ', '').replace('　', '')
            if '施術の種類' in v or 'はり・きゅう' in v or '１術' in v:
                for co in range(1, spot_ws.max_column + 1):
                    cv = str(spot_ws.cell(r, co).value or '').replace(' ', '').replace('　', '')
                    if '１術' in cv or '1術' in cv:
                        for c_next in range(co + 1, co + 8):
                            cnt_v = spot_ws.cell(r, c_next).value
                            if cnt_v and str(cnt_v).strip() and '回' in str(cnt_v):
                                jutsu1 = str(cnt_v).strip()
                                break
                    elif '２術' in cv or '2術' in cv:
                        for c_next in range(co + 1, co + 8):
                            cnt_v = spot_ws.cell(r, c_next).value
                            if cnt_v and str(cnt_v).strip() and '回' in str(cnt_v):
                                jutsu2 = str(cnt_v).strip()
                                break
                # 摘要欄 (BK列 = column 63)
                for r_sub in range(r, r + 6):
                    s_val = spot_ws.cell(r_sub, 63).value
                    if s_val and str(s_val).strip() and str(s_val).strip().replace(' ', '').replace('　', '') != '摘要':
                        summary_text = str(s_val).strip()
                        break
                                
    return {'jutsu1': jutsu1, 'jutsu2': jutsu2, 'summary': summary_text}


def extract_period_and_days_dynamic(spot_ws):
    """初療年月日、施術期間（自・至）、実日数を動的抽出"""
    shoryou_date, from_date, to_date, actual_days = None, None, None, None
    r_header = None
    for r in range(40, 70):
        for c in range(1, 15):
            v = str(spot_ws.cell(r, c).value or '').replace(' ', '').replace('　', '')
            if '初療年月日' in v or '初療' in v:
                r_header = r
                break
        if r_header: break
        
    if r_header:
        for r in range(r_header + 1, r_header + 7):
            for c in range(1, spot_ws.max_column + 1):
                raw_v = spot_ws.cell(r, c).value
                if raw_v is None: continue
                s_v = str(raw_v).strip()
                if ('令和' in s_v or '平成' in s_v) and '年' in s_v:
                    if c < 20 and not shoryou_date:
                        shoryou_date = s_v
                    elif c < 35 and not from_date:
                        from_date = s_v
                    elif c >= 35 and not to_date:
                        to_date = s_v
                if isinstance(raw_v, (int, float)) and 1 <= raw_v <= 31 and c >= 45 and not actual_days:
                    actual_days = int(raw_v)
                elif s_v.isdigit() and 1 <= int(s_v) <= 31 and c >= 45 and not actual_days:
                    actual_days = int(s_v)
                    
    return {
        'shoryou': shoryou_date,
        'from_date': from_date,
        'to_date': to_date,
        'actual_days': actual_days
    }


def extract_payment_account_info_dynamic(spot_ws):
    """支払機関欄の口座番号（数字配列）および口座名義（カタカナ）を動的抽出"""
    acc_digits = []
    holder_name = None
    for r in range(190, min(260, spot_ws.max_row + 1)):
        for c in range(1, 20):
            v = str(spot_ws.cell(r, c).value or '').replace(' ', '').replace('　', '')
            if '口座名義' in v or 'カタカナで記入' in v or '口座番号' in v:
                for co in range(c + 1, 35):
                    val = spot_ws.cell(r, co).value
                    if val and str(val).strip() and str(val).strip() not in ['口座名義', '口座番号'] and not holder_name:
                        holder_name = str(val).strip()
                for r_d in range(r, r + 2):
                    for co in range(35, spot_ws.max_column + 1):
                        cv = spot_ws.cell(r_d, co).value
                        if cv is not None and str(cv).strip().isdigit():
                            acc_digits.append(str(cv).strip())
                if holder_name or acc_digits: break
        if holder_name or acc_digits: break
    return {'digits': acc_digits, 'holder': holder_name}


def extract_kikan_code_dynamic(spot_ws):
    """機関コード（保険医療機関コード）を動的抽出"""
    for r in range(4, 15):
        for c in range(35, spot_ws.max_column + 1):
            v = str(spot_ws.cell(r, c).value or '').replace(' ', '').replace('　', '')
            if '機関コード' in v:
                for co in range(c + 1, c + 10):
                    val = spot_ws.cell(r, co).value
                    if val and str(val).strip() and str(val).strip() != '機関コード':
                        return str(val).strip()
            elif re.search(r'^\d{3}-\d{3}$', v) or re.search(r'^\d{7,10}$', v) and c >= 45:
                return v
    return spot_ws['AY8'].value or ""


def extract_claim_year_month_dynamic(spot_ws):
    """請求年月（令和○年○月分）を動的抽出"""
    for r in range(1, 10):
        for c in range(1, spot_ws.max_column + 1):
            v = str(spot_ws.cell(r, c).value or '').strip()
            if ('令和' in v or '平成' in v) and '年' in v and '月' in v:
                return v
    return spot_ws["AH4"].value or spot_ws["E5"].value


def extract_treatment_location(ws):
    """「施術した場所」の記載内容をspotlogから抽出（通常時は請求区分見出しなどを拾わずNone）"""
    for r in range(44, 50):
        for c in range(40, 75):
            v = ws.cell(row=r, column=c).value
            if v is not None and str(v).strip() != '':
                str_v = str(v).strip()
                clean_v = re.sub(r'[\s　]+', '', str_v)
                if not any(k in clean_v for k in ['施術した場所', '実日数', '請求区分', '新規', '継続', '初療', '期間']):
                    return str_v
    return None


def extract_consent_record(ws):
    """同意記録（同意医師の氏名、住所、同意年月日、傷病名、要加療期間）を動的抽出"""
    header_row = None
    for r in range(200, 240):
        for c in range(1, 10):
            v = str(ws.cell(row=r, column=c).value or '')
            if '同意医師' in v or '同意\n記録' in v or '同意記録' in v:
                header_row = r
                break
        if header_row:
            break
            
    if not header_row:
        return None
        
    for r in range(header_row + 1, header_row + 7):
        for c in range(1, 15):
            v = ws.cell(row=r, column=c).value
            if v is not None and str(v).strip() != '' and str(v).strip() != '同意医師の氏名':
                doc_name = str(v).strip()
                addr = str(ws.cell(row=r, column=19).value or ws.cell(row=r, column=18).value or '')
                c_date = str(ws.cell(row=r, column=44).value or ws.cell(row=r, column=43).value or '')
                disease = str(ws.cell(row=r, column=57).value or ws.cell(row=r, column=56).value or '')
                period = str(ws.cell(row=r, column=70).value or ws.cell(row=r, column=69).value or '')
                return {
                    'doc_name': doc_name,
                    'address': addr,
                    'consent_date': c_date,
                    'disease': disease,
                    'period': period
                }
    return None


def extract_birthdate_formatted(ws):
    """患者の生年月日をspotlogから動的に抽出し、年号丸囲み付きで整形"""
    for r in range(40, 55):
        for c in range(1, 35):
            v = ws.cell(row=r, column=c).value
            if v:
                str_v = str(v).strip()
                m = re.search(r'(明治|大正|昭和|平成|令和|明|大|昭|平|令)\s*([0-9０-９]+|元)\s*年\s*([0-9０-９]+)\s*月\s*([0-9０-９]+)\s*日', str_v)
                if m:
                    era_raw = m.group(1)
                    y = m.group(2)
                    m_val = m.group(3)
                    d = m.group(4)
                    
                    if '明' in era_raw: era_str = '○明・大・昭・平・令'
                    elif '大' in era_raw: era_str = '明・○大・昭・平・令'
                    elif '昭' in era_raw: era_str = '明・大・○昭・平・令'
                    elif '平' in era_raw: era_str = '明・大・昭・○平・令'
                    elif '令' in era_raw: era_str = '明・大・昭・平・○令'
                    else: era_str = '明・大・昭・平・令'
                    
                    return f"{era_str}　{y}年　{m_val}月　{d}日生"
                    
    return "明・大・昭・平・令　　年　　月　　日生"


def detect_spotlog_type(ws):
    """
    シートの右上の様式番号を確認してspotlog様式か判定:
    - 「様式第５号の３」 (はり・きゅう用 spotlog) -> '鍼灸'
    - 「様式第５号の４」 (あんま・マッサージ用 spotlog) -> 'マッサージ'
    - 「様式第５号の１」 / 「様式第５号の２」 (すでに基準様式) または その他形式 -> None (スキップ)
    """
    for r in range(1, 6):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v:
                str_v = str(v).replace(' ', '').replace('　', '')
                str_v = str_v.replace('５', '5').replace('３', '3').replace('４', '4').replace('１', '1').replace('２', '2')
                if "様式第5号の3" in str_v or "第5号の3" in str_v or "5号の3" in str_v:
                    return "鍼灸"
                elif "様式第5号の4" in str_v or "第5号の4" in str_v or "5号の4" in str_v:
                    return "マッサージ"
                elif "様式第5号の1" in str_v or "様式第5号の2" in str_v:
                    return None
    return None


def clone_worksheet_to_wb(src_ws, dest_wb, new_title):
    """別ワークブックからシートのレイアウト・フォント・罫線・列幅(min/max範囲含む)・行高・結合を100%完全複製"""
    dest_ws = dest_wb.create_sheet(title=new_title)
    
    dest_ws.sheet_format = copy(src_ws.sheet_format)
    dest_ws.sheet_properties = copy(src_ws.sheet_properties)
    dest_ws.page_setup = copy(src_ws.page_setup)
    dest_ws.print_options = copy(src_ws.print_options)
    dest_ws.page_margins = copy(src_ws.page_margins)
    
    for r, rd in src_ws.row_dimensions.items():
        dest_rd = dest_ws.row_dimensions[r]
        dest_rd.height = rd.height
        dest_rd.hidden = rd.hidden
        dest_rd.outline_level = rd.outline_level
            
    for k, v in src_ws.column_dimensions.items():
        cd = openpyxl.worksheet.dimensions.ColumnDimension(
            dest_ws, index=k, min=v.min, max=v.max, width=v.width, hidden=v.hidden, outline_level=v.outline_level
        )
        dest_ws.column_dimensions[k] = cd
        
    for rng in src_ws.merged_cells.ranges:
        dest_ws.merge_cells(str(rng))
        
    for row in src_ws.iter_rows():
        for cell in row:
            if cell.value is not None or cell.has_style:
                new_cell = dest_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = copy(cell.number_format)
                    new_cell.protection = copy(cell.protection)
                    new_cell.alignment = copy(cell.alignment)
    return dest_ws


def convert_acupuncture_dynamic(spot_ws, target_ws):
    """はり・きゅう用の完全正確配置転記（全行の「円×」「回＝」「円」列位置を完全垂直一致）"""
    img_coords = get_image_anchors(spot_ws)
    
    # 1. タイトル年月（動的抽出）
    val_ym = extract_claim_year_month_dynamic(spot_ws)
    month_num = "7"
    if val_ym:
        str_ym = str(val_ym)
        target_ws["E5"] = f"療 養 費 支 給 申 請 書{val_ym}（はり・きゅう用）" if "（" in str_ym else f"療 養 費 支 給 申 請 書（{val_ym}）（はり・きゅう用）"
        m_match = re.search(r'(\d+)\s*月', str_ym)
        if m_match:
            month_num = m_match.group(1)

    # カレンダーの「月」枠（縦書き 7\n月）
    target_ws["X164"] = f"{month_num}\n月"
    target_ws["X164"].alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    target_ws["X164"].font = Font(name="ＭＳ 明朝", size=10, bold=True)

    # 機関コード (CC9 動的抽出)
    k_code = extract_kikan_code_dynamic(spot_ws)
    if k_code:
        target_ws["CC9"] = f"機関コード　{k_code}"

    # 種類 (CY24: 05 鍼灸)
    target_ws["CY24"] = "05 鍼灸"

    # 公費負担者番号 (Row 13: 8マス)
    fill_header_boxes(target_ws, 13, extract_header_box_digits(spot_ws, 12))
    # 公費受給者番号 (Row 18: 7〜8マス)
    fill_header_boxes(target_ws, 18, extract_header_box_digits(spot_ws, 16))
    # 区市町村番号 (Row 23: 6〜8マス)
    fill_header_boxes(target_ws, 23, extract_header_box_digits(spot_ws, 20))
    # 受給者番号 (Row 28: 7〜8マス)
    fill_header_boxes(target_ws, 28, extract_header_box_digits(spot_ws, 24))

    # 特記事項・保険種別・給付割合 (完全動的丸囲み)
    marks = detect_special_marks(spot_ws, img_coords)
    target_ws["CU15"] = marks['shakoku']
    target_ws["DD15"] = marks['koukou']
    target_ws["CU19"] = marks['kouhi']
    target_ws["DM13"] = marks['hon_gai']
    target_ws["DM16"] = marks['roku_gai']
    target_ws["DM19"] = marks['ie_gai']
    target_ws["DU15"] = marks['kou_gai_1']
    target_ws["DU18"] = marks['kou_gai_7']
    target_ws["EG18"] = marks['rate_8']
    target_ws["EM18"] = marks['rate_9']
    target_ws["ES18"] = marks['rate_10']

    # 1. ヘッダー・被保険者欄（完全動的抽出）
    h_data = extract_header_data_dynamic(spot_ws)

    # 保険者番号 (8桁個別ボックス: 右詰めで配置)
    box_cols = ["DC28", "DI28", "DO28", "DU28", "EA28", "EG28", "EM28", "ES28"]
    for col in box_cols:
        target_ws[col] = None
    if h_data['ins_digits']:
        digs = h_data['ins_digits'][-8:] if len(h_data['ins_digits']) > 8 else h_data['ins_digits']
        start_idx = len(box_cols) - len(digs)
        for i, d in enumerate(digs):
            target_ws[box_cols[start_idx + i]] = str(d)

    # 被保険者記号番号, 発病年月日, 原因経過
    if h_data['kigou']:
        target_ws["J38"] = str(h_data['kigou'])
    if h_data['hatsubyou']:
        target_ws["AU38"] = str(h_data['hatsubyou'])
        target_ws["AU38"].alignment = Alignment(horizontal="center", vertical="center")
    if h_data['genin']:
        clean_cause = re.sub(r'[\r\n]+', '', str(h_data['genin'])).strip()
        target_ws["BX38"] = clean_cause
        target_ws["BX38"].alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
        target_ws["BX38"].font = Font(name="ＭＳ 明朝", size=9)
        
    # フリガナ, 氏名, 性別, 続柄
    if h_data['kana']:
        target_ws["AC43"] = f"(ﾌﾘｶﾞﾅ) {h_data['kana']}"
    if h_data['name']:
        target_ws["AC47"] = str(h_data['name'])
    if h_data['gender']:
        target_ws["BQ47"] = "○男・女" if h_data['gender'] == "男" else "男・○女"
    if h_data['relation']:
        target_ws["BW47"] = str(h_data['relation'])
    
    # 施術した場所 (CJ52 & CJ57: 長文でもはみ出さない自動縮小＋折り返し)
    loc_text = extract_treatment_location(spot_ws)
    if loc_text:
        target_ws["CJ52"] = "○施術した場所（施設等に入居している場合及び被保険者の住所と異なる場合に記載）"
        target_ws["CJ57"] = loc_text
        try:
            target_ws.merge_cells("CJ57:EX61")
        except Exception:
            pass
            
        f_size = 7.0 if len(loc_text) > 70 else (8.0 if len(loc_text) > 40 else 9.0)
        target_ws["CJ57"].alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
        target_ws["CJ57"].font = Font(name="ＭＳ 明朝", size=f_size)
    else:
        target_ws["CJ52"] = "○施術した場所（施設等に入居している場合及び被保険者の住所と異なる場合に記載）"
        target_ws["CJ57"] = None
    
    # 生年月日 (動的抽出)
    target_ws["AC57"] = extract_birthdate_formatted(spot_ws)
    
    # 業務上・外・第三者行為 (動的判定)
    target_ws["CJ47"] = detect_work_injury(spot_ws, img_coords)

    # 施術内容欄（動的抽出）
    p_days = extract_period_and_days_dynamic(spot_ws)
    if p_days.get('shoryou'): target_ws["J67"] = str(p_days['shoryou'])
    target_ws["AQ67"] = f"自・{p_days.get('from_date') or ''} ～至・{p_days.get('to_date') or ''}"
    if p_days.get('actual_days'): target_ws["DC67"] = f"{p_days['actual_days']}日"
    
    # 請求区分 & 転帰 (動的判定)
    target_ws["DO67"] = detect_claim_type(spot_ws, img_coords)
    target_ws["DO67"].alignment = Alignment(horizontal="center", vertical="center")
    target_ws["DO77"] = detect_outcome(spot_ws, img_coords)
    target_ws["DO77"].alignment = Alignment(horizontal="center", vertical="center")

    # 傷病名（動的判定）
    dis_res = detect_acupuncture_diseases(spot_ws, img_coords)
    target_ws["AE72"] = dis_res['d1']
    target_ws["AX72"] = dis_res['d2']
    target_ws["BW72"] = dis_res['d3']
    target_ws["CS72"] = dis_res['d4']
    target_ws["AE77"] = dis_res['d5']
    target_ws["AX77"] = dis_res['d6']
    target_ws["BW77"] = dis_res['d7']

    # 初検料（動的判定）
    fe_res = detect_first_exam_fee(spot_ws, img_coords)
    target_ws["J82"] = fe_res['text']
    if fe_res['price']:
        target_ws["BF82"] = f"{fe_res['price']} 円"

    # 施術の種類 (1術・2術 回数 & 摘要欄 動的抽出)
    ac_jutsu = extract_acupuncture_jutsu_counts_dynamic(spot_ws)
    if ac_jutsu.get('jutsu1'): target_ws["BW87"] = f"１術 {ac_jutsu['jutsu1']}"
    if ac_jutsu.get('jutsu2'): target_ws["CT87"] = f"２術 {ac_jutsu['jutsu2']}"

    # 中部：各施術料・加算・合計・一部負担金・請求額（完全動的抽出）
    m_fees = extract_middle_fees_dynamic(spot_ws)

    # 通所 (Row 92..99)
    if 'tuusho' in m_fees:
        tf = m_fees['tuusho']
        safe_apply_grid_row(target_ws, 92, 99, tf['u_price'], tf['u_count'], tf['u_total'], tf['l_price'], tf['l_count'], tf['l_total'], is_boxed=True)

    # 訪問施術料１ (Row 100..107)
    if 'h1' in m_fees:
        tf = m_fees['h1']
        safe_apply_grid_row(target_ws, 100, 107, tf['u_price'], tf['u_count'], tf['u_total'], tf['l_price'], tf['l_count'], tf['l_total'], is_boxed=False)

    # 訪問施術料２ (Row 108..115)
    if 'h2' in m_fees:
        tf = m_fees['h2']
        safe_apply_grid_row(target_ws, 108, 115, tf['u_price'], tf['u_count'], tf['u_total'], tf['l_price'], tf['l_count'], tf['l_total'], is_boxed=False)

    # 訪問施術料３ (Row 116..123)
    if 'h3' in m_fees:
        tf = m_fees['h3']
        safe_apply_grid_row(target_ws, 116, 123, tf['u_price'], tf['u_count'], tf['u_total'], tf['l_price'], tf['l_count'], tf['l_total'], is_boxed=False)

    # 電療料 (Row 124..128)
    target_ws["O124"] = detect_electrotherapy(spot_ws, img_coords)
    if 'denryou' in m_fees:
        tf = m_fees['denryou']
        safe_apply_grid_row(target_ws, 124, 128, tf['price'], tf['count'], tf['total'], is_boxed=False)

    # 特別地域 (Row 129..133)
    if 'tokubetsu' in m_fees:
        tf = m_fees['tokubetsu']
        safe_apply_grid_row(target_ws, 129, 133, tf['price'], tf['count'], tf['total'], is_boxed=False)

    # 往療料 (Row 134..138)
    if 'ouryou' in m_fees:
        tf = m_fees['ouryou']
        safe_apply_grid_row(target_ws, 134, 138, tf['price'], tf['count'], tf['total'], is_boxed=False)

    # 施術報告書交付料 (Row 139..143)
    prev_ym_hari = extract_report_prev_date(spot_ws, is_massage=False)
    target_ws["J139"] = f"施術報告書交付料（前回支給：{prev_ym_hari}）"
    if 'houkoku' in m_fees:
        tf = m_fees['houkoku']
        safe_apply_grid_row(target_ws, 139, 143, tf['price'], tf['count'], tf['total'], is_boxed=False)

    # 明細書発行加算 (Row 144..148)
    if 'meisai' in m_fees and m_fees['meisai']:
        tf = m_fees['meisai']
        safe_apply_grid_row(target_ws, 144, 148, tf['price'], tf['count'], tf['total'], is_boxed=False)

    # 合計
    if 'goukei' in m_fees and m_fees['goukei'] is not None:
        target_ws["BF149"] = f"{format_currency_str(m_fees['goukei'])} 円"
    else:
        target_ws["BF149"] = None
        
    # 一部負担金 & 請求額
    target_ws["J154"] = extract_copayment_ratio_text(spot_ws, marks)
    if 'ichibu' in m_fees and m_fees['ichibu'] is not None:
        target_ws["BF154"] = f"{format_currency_str(m_fees['ichibu'])} 円"
    else:
        target_ws["BF154"] = None
        
    if 'seikyuu' in m_fees and m_fees['seikyuu'] is not None:
        target_ws["BF159"] = f"{format_currency_str(m_fees['seikyuu'])} 円"
    else:
        target_ws["BF159"] = None

    # 摘要欄 (DO87:EX163 を結合して全文表示)
    try:
        target_ws.merge_cells("DO87:EX163")
    except Exception:
        pass
    summary_val = ac_jutsu.get('summary') or spot_ws["BK73"].value
    if summary_val:
        target_ws["DO87"] = str(summary_val)
        target_ws["DO87"].alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
        target_ws["DO87"].font = Font(name="ＭＳ 明朝", size=9)

    # カレンダー (Row 168: 中央揃えで動的配置)
    cal_data = extract_calendar_marks_dynamic(spot_ws)
    for d, mark in cal_data.items():
        if 1 <= d <= 31:
            cal_col_idx = 30 + (d - 1) * 4
            cell = target_ws.cell(row=168, column=cal_col_idx, value=str(mark))
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(name="ＭＳ 明朝", size=11, bold=True)

    # 往療又は訪問の理由 (動的複数選択対応)
    target_ws["J173"] = detect_visit_reasons(spot_ws, img_coords)

    # 施術証明欄 (完全動的抽出)
    c_sec = extract_cert_section_dynamic(spot_ws)
    target_ws["DC177"] = detect_practitioner_location_type(spot_ws, img_coords)
    if c_sec.get('zip'): target_ws["CR181"] = f"〒{str(c_sec['zip']).replace('〒', '')}"
    if c_sec.get('date'): target_ws["M184"] = str(c_sec['date'])
    if c_sec.get('addr'): target_ws["CR184"] = str(c_sec['addr'])
    if c_sec.get('reg_no'): target_ws["M191"] = str(c_sec['reg_no'])
    
    try:
        target_ws.merge_cells("CR188:EX191")
    except Exception:
        pass
    if c_sec.get('clinic_name'):
        target_ws["CR188"] = str(c_sec['clinic_name'])
        target_ws["CR188"].alignment = Alignment(vertical="center", horizontal="left")
        target_ws["CR188"].font = Font(name="ＭＳ 明朝", size=10)
        
    if c_sec.get('manager_name'): target_ws["CR192"] = str(c_sec['manager_name'])
    if c_sec.get('tel'): target_ws["EG192"] = str(c_sec['tel'])

    # 申請欄 (完全動的抽出)
    a_sec = extract_applicant_section_dynamic(spot_ws)
    if a_sec.get('zip'): target_ws["CR196"] = f"〒{str(a_sec['zip']).replace('〒', '')}"
    if a_sec.get('date'): target_ws["M200"] = str(a_sec['date'])
    if a_sec.get('addr'): target_ws["CR200"] = str(a_sec['addr'])
    if a_sec.get('kouiki'): target_ws["M204"] = str(a_sec['kouiki'])
    if a_sec.get('name'): target_ws["CR207"] = str(a_sec['name'])
    if a_sec.get('tel'): target_ws["EG207"] = str(a_sec['tel'])

    # 支払機関欄 (完全動的判定)
    pay_sec = detect_payment_section(spot_ws, img_coords)
    
    target_ws["L214"] = "①．" if pay_sec['pay_1'] else "1．"
    target_ws["AH214"] = "②．" if pay_sec['pay_2'] else "2．"
    target_ws["L217"] = "③．" if pay_sec['pay_3'] else "3．"
    target_ws["AH217"] = "④．" if pay_sec['pay_4'] else "4．"
    
    target_ws["BD214"] = "①.　普通" if pay_sec['dep_1'] else "1.　普通"
    target_ws["BS214"] = "②.　当座" if pay_sec['dep_2'] else "2.　当座"
    target_ws["BD217"] = "③.　通知" if pay_sec['dep_3'] else "3.　通知"
    target_ws["BS217"] = "④.　別段" if pay_sec['dep_4'] else "4.　別段"
    
    # 金融機関名（右側の入力枠 CW211:DQ219 に中央揃えで配置、長い名前も自動文字縮小対応）
    b_name = pay_sec['bank_name']
    try:
        target_ws.merge_cells("CW211:DQ219")
    except Exception:
        pass
    target_ws["CW211"] = b_name
    b_font_size = 7.5 if len(b_name) > 10 else (8.5 if len(b_name) > 6 else 10.0)
    target_ws["CW211"].font = Font(name="ＭＳ 明朝", size=b_font_size)
    target_ws["CW211"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    b_t = pay_sec['bank_type']
    target_ws["DR211"] = "○銀行" if "銀行" in b_t else "銀行"
    target_ws["DR214"] = "○金庫" if any(k in b_t for k in ["金庫", "信金"]) else "金庫"
    target_ws["DR217"] = "○農協" if any(k in b_t for k in ["農協", "JA"]) else "農協"
    
    # 支店名（DY211:EN219 に配置、長い名前も自動文字縮小対応）
    br_name = pay_sec['branch_name']
    try:
        target_ws.merge_cells("DY211:EN219")
    except Exception:
        pass
    target_ws["DY211"] = br_name
    br_font_size = 7.5 if len(br_name) > 10 else (8.5 if len(br_name) > 6 else 10.0)
    target_ws["DY211"].font = Font(name="ＭＳ 明朝", size=br_font_size)
    target_ws["DY211"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    br_t = pay_sec['branch_type']
    target_ws["EO211"] = "○本店" if "本店" in br_t else "本店"
    target_ws["EO214"] = "○支店" if "支店" in br_t else "支店"
    target_ws["EO217"] = "○出張所" if "出張所" in br_t else "出張所"
    
    # 口座番号 & 口座名義（完全動的抽出: 右詰めで配置）
    acc_boxes = ["CG220", "CL220", "CQ220", "CV220", "DA220", "DF220", "DK220", "DP220"]
    for col in acc_boxes:
        target_ws[col] = None
        
    p_acc = extract_payment_account_info_dynamic(spot_ws)
    acc_digits = p_acc['digits']
    if acc_digits:
        digs = acc_digits[-8:] if len(acc_digits) > 8 else acc_digits
        start_idx = len(acc_boxes) - len(digs)
        for i, d in enumerate(digs):
            target_ws[acc_boxes[start_idx + i]] = str(d)
        
    if p_acc.get('holder'):
        target_ws["AD220"] = str(p_acc['holder'])

    # 同意記録 (動的探索)
    c_rec = extract_consent_record(spot_ws)
    if c_rec:
        target_ws["J229"] = c_rec['doc_name']
        target_ws["AJ229"] = c_rec['address']
        target_ws["CA229"] = c_rec['consent_date']
        target_ws["DE229"] = c_rec['disease']
        target_ws["EF229"] = c_rec['period']
    else:
        target_ws["J229"] = None
        target_ws["AJ229"] = None
        target_ws["DE229"] = None
        target_ws["EF229"] = None

    # 委任状欄 (完全動的抽出)
    d_sec = extract_delegation_section_dynamic(spot_ws)
    if d_sec.get('date'): target_ws["CP240"] = str(d_sec['date'])
    
    target_ws["J244"] = "申請者"
    target_ws["J244"].alignment = Alignment(horizontal="center", vertical="center")
    
    target_ws["CD244"] = "代理人"
    target_ws["CD244"].alignment = Alignment(horizontal="center", vertical="center")
    
    target_ws["J252"] = "（被保険者）"
    target_ws["J252"].alignment = Alignment(horizontal="center", vertical="center")
    
    target_ws["CD252"] = None
    
    applicant_addr = str(d_sec.get('applicant_addr') or "")
    target_ws["AA244"] = f"住所　{applicant_addr}" if applicant_addr else "住所　"
    target_ws["AA244"].alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
    target_ws["AA244"].font = Font(name="ＭＳ 明朝", size=10)
    
    target_ws["CU244"] = format_proxy_address(d_sec.get('delegate_addr'))
    target_ws["CU244"].alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
    target_ws["CU244"].font = Font(name="ＭＳ 明朝", size=8.5)
    
    applicant_name = str(d_sec.get('applicant_name') or "")
    target_ws["AA252"] = f"氏名　{applicant_name}" if applicant_name else "氏名　"
    target_ws["AA252"].alignment = Alignment(vertical="center", horizontal="left")
    target_ws["AA252"].font = Font(name="ＭＳ 明朝", size=10)
    
    proxy_name = str(d_sec.get('delegate_name') or "")
    target_ws["CU252"] = f"氏名　{proxy_name}" if proxy_name else "氏名　"
    target_ws["CU252"].alignment = Alignment(vertical="center", horizontal="left")
    target_ws["CU252"].font = Font(name="ＭＳ 明朝", size=10)

    patient_name = h_data.get('name') or spot_ws["O40"].value or spot_ws.title
    return str(patient_name).strip()


def convert_massage_dynamic(spot_ws, target_ws):
    """あんま・マッサージ用の完全正確配置転記（全行の「円×」「回＝」「円」列位置を完全垂直一致）"""
    img_coords = get_image_anchors(spot_ws)
    
    # 1. タイトル年月（動的抽出）
    val_ym = extract_claim_year_month_dynamic(spot_ws)
    month_num = "7"
    if val_ym:
        str_ym = str(val_ym)
        target_ws["E5"] = f"療 養 費 支 給 申 請 書{val_ym}（あんま・マッサージ用）" if "（" in str_ym else f"療 養 費 支 給 申 請 書（{val_ym}）（あんま・マッサージ用）"
        m_match = re.search(r'(\d+)\s*月', str_ym)
        if m_match:
            month_num = m_match.group(1)

    # カレンダーの「月」枠（縦書き 7\n月）
    target_ws["X168"] = f"{month_num}\n月"
    target_ws["X168"].alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    target_ws["X168"].font = Font(name="ＭＳ 明朝", size=10, bold=True)

    # 機関コード (CC9 動的抽出)
    k_code = extract_kikan_code_dynamic(spot_ws)
    if k_code:
        target_ws["CC9"] = f"機関コード　{k_code}"

    # 公費負担者番号 (Row 13: 8マス)
    fill_header_boxes(target_ws, 13, extract_header_box_digits(spot_ws, 12))
    # 公費受給者番号 (Row 18: 7〜8マス)
    fill_header_boxes(target_ws, 18, extract_header_box_digits(spot_ws, 16))
    # 区市町村番号 (Row 23: 6〜8マス)
    fill_header_boxes(target_ws, 23, extract_header_box_digits(spot_ws, 20))
    # 受給者番号 (Row 28: 7〜8マス)
    fill_header_boxes(target_ws, 28, extract_header_box_digits(spot_ws, 24))

    # 特記事項・保険種別・給付割合 (完全動的丸囲み)
    marks = detect_special_marks(spot_ws, img_coords)
    target_ws["CU15"] = marks['shakoku']
    target_ws["DD15"] = marks['koukou']
    target_ws["CU19"] = marks['kouhi']
    target_ws["DM13"] = marks['hon_gai']
    target_ws["DM16"] = marks['roku_gai']
    target_ws["DM19"] = marks['ie_gai']
    target_ws["DU15"] = marks['kou_gai_1']
    target_ws["DU18"] = marks['kou_gai_7']
    target_ws["EG18"] = marks['rate_8']
    target_ws["EM18"] = marks['rate_9']
    target_ws["ES18"] = marks['rate_10']

    # 1. ヘッダー・被保険者欄（完全動的抽出）
    h_data = extract_header_data_dynamic(spot_ws)

    # 保険者番号 (8桁個別ボックス: 右詰めで配置)
    box_cols = ["DC28", "DI28", "DO28", "DU28", "EA28", "EG28", "EM28", "ES28"]
    for col in box_cols:
        target_ws[col] = None
    if h_data['ins_digits']:
        digs = h_data['ins_digits'][-8:] if len(h_data['ins_digits']) > 8 else h_data['ins_digits']
        start_idx = len(box_cols) - len(digs)
        for i, d in enumerate(digs):
            target_ws[box_cols[start_idx + i]] = str(d)

    # 被保険者記号番号, 発病年月日, 原因経過
    if h_data['kigou']:
        target_ws["J38"] = str(h_data['kigou'])
    if h_data['hatsubyou']:
        target_ws["AU38"] = str(h_data['hatsubyou'])
        target_ws["AU38"].alignment = Alignment(horizontal="center", vertical="center")
    if h_data['genin']:
        clean_cause = re.sub(r'[\r\n]+', '', str(h_data['genin'])).strip()
        target_ws["BX38"] = clean_cause
        target_ws["BX38"].alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
        target_ws["BX38"].font = Font(name="ＭＳ 明朝", size=9)
        
    # フリガナ, 氏名, 性別, 続柄
    if h_data['kana']:
        target_ws["AC43"] = f"(ﾌﾘｶﾞﾅ) {h_data['kana']}"
    if h_data['name']:
        target_ws["AC47"] = str(h_data['name'])
    if h_data['gender']:
        target_ws["BQ47"] = "○男・女" if h_data['gender'] == "男" else "男・○女"
    if h_data['relation']:
        target_ws["BW47"] = str(h_data['relation'])
    
    # 施術した場所 (CJ52 & CJ57: 長文でもはみ出さない自動縮小＋折り返し)
    loc_text = extract_treatment_location(spot_ws)
    if loc_text:
        target_ws["CJ52"] = "○施術した場所（施設等に入居している場合及び被保険者の住所と異なる場合記載）"
        target_ws["CJ57"] = loc_text
        try:
            target_ws.merge_cells("CJ57:EX61")
        except Exception:
            pass
            
        f_size = 7.0 if len(loc_text) > 70 else (8.0 if len(loc_text) > 40 else 9.0)
        target_ws["CJ57"].alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
        target_ws["CJ57"].font = Font(name="ＭＳ 明朝", size=f_size)
    else:
        target_ws["CJ52"] = "○施術した場所（施設等に入居している場合及び被保険者の住所と異なる場合記載）"
        target_ws["CJ57"] = None
    
    # 生年月日 (動的抽出)
    target_ws["AC57"] = extract_birthdate_formatted(spot_ws)
    
    # 業務上・外・第三者行為 (動的判定)
    target_ws["CJ47"] = detect_work_injury(spot_ws, img_coords)

    # 施術内容欄（動的抽出）
    p_days = extract_period_and_days_dynamic(spot_ws)
    if p_days.get('shoryou'): target_ws["J67"] = str(p_days['shoryou'])
    target_ws["AQ67"] = f"自・{p_days.get('from_date') or ''} ～至・{p_days.get('to_date') or ''}"
    if p_days.get('actual_days'): target_ws["DC67"] = f"{p_days['actual_days']}日"
    
    # 請求区分 & 転帰 (動的判定)
    target_ws["DO67"] = detect_claim_type(spot_ws, img_coords)
    target_ws["DO67"].alignment = Alignment(horizontal="center", vertical="center")
    target_ws["DO76"] = detect_outcome(spot_ws, img_coords)
    target_ws["DO76"].alignment = Alignment(horizontal="center", vertical="center")

    # 傷病名及び症状 (動的検索)
    diag_text = ""
    for r in range(40, 80):
        for c in range(1, 15):
            v = str(spot_ws.cell(r, c).value or '').replace(' ', '').replace('　', '')
            if '傷病名及び症状' in v or '傷病名' in v:
                for co in range(c + 1, spot_ws.max_column + 1):
                    dv = spot_ws.cell(r, co).value
                    if dv and len(str(dv).strip()) > 3 and not diag_text:
                        diag_text = str(dv).strip()
                        break
            if diag_text: break
        if diag_text: break
    if not diag_text:
        diag_text = str(spot_ws["W59"].value or "").strip()
    target_ws["AQ72"] = diag_text
    target_ws["AQ72"].alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
    target_ws["AQ72"].font = Font(name="ＭＳ Ｐ明朝", size=8.0)

    # 同意部位・施術回数 (Row 84 動的抽出)
    mb_counts = extract_massage_body_counts_dynamic(spot_ws)
    target_ws["BQ84"] = mb_counts.get('kukan', '')
    target_ws["BQ84"].alignment = Alignment(horizontal="center", vertical="center")
    target_ws["CA84"] = mb_counts.get('r_arm', '')
    target_ws["CA84"].alignment = Alignment(horizontal="center", vertical="center")
    target_ws["CK84"] = mb_counts.get('l_arm', '')
    target_ws["CK84"].alignment = Alignment(horizontal="center", vertical="center")
    target_ws["CU84"] = mb_counts.get('r_leg', '')
    target_ws["CU84"].alignment = Alignment(horizontal="center", vertical="center")
    target_ws["DE84"] = mb_counts.get('l_leg', '')
    target_ws["DE84"].alignment = Alignment(horizontal="center", vertical="center")

    # 中部：各施術料・加算・合計・一部負担金・請求額（完全動的抽出）
    m_fees = extract_middle_fees_dynamic(spot_ws)

    # 通所 (Row 88..95)
    if 'tuusho' in m_fees:
        tf = m_fees['tuusho']
        safe_apply_grid_row(target_ws, 88, 95, tf['u_price'], tf['u_count'], tf['u_total'], tf['l_price'], tf['l_count'], tf['l_total'], is_boxed=False)

    # 訪問施術料１ (Row 96..103)
    if 'h1' in m_fees:
        tf = m_fees['h1']
        safe_apply_grid_row(target_ws, 96, 103, tf['u_price'], tf['u_count'], tf['u_total'], tf['l_price'], tf['l_count'], tf['l_total'], is_boxed=False)

    # 訪問施術料２ (Row 104..111)
    if 'h2' in m_fees:
        tf = m_fees['h2']
        safe_apply_grid_row(target_ws, 104, 111, tf['u_price'], tf['u_count'], tf['u_total'], tf['l_price'], tf['l_count'], tf['l_total'], is_boxed=False)

    # 訪問施術料３ (Row 112..119)
    if 'h3' in m_fees:
        tf = m_fees['h3']
        safe_apply_grid_row(target_ws, 112, 119, tf['u_price'], tf['u_count'], tf['u_total'], tf['l_price'], tf['l_count'], tf['l_total'], is_boxed=False)

    # 温罨法 (加算 Row 120..123)
    if 'on_an' in m_fees:
        tf = m_fees['on_an']
        safe_apply_grid_row(target_ws, 120, 123, tf['price'], tf['count'], tf['total'], is_boxed=False)

    # 温罨法・電気光線器具 (加算 Row 124..127)
    if 'on_den' in m_fees:
        tf = m_fees['on_den']
        safe_apply_grid_row(target_ws, 124, 127, tf['price'], tf['count'], tf['total'], is_boxed=False)

    # 変形徒手矯正術 (Row 132 & Row 136..139 動的抽出)
    target_ws["BS132"] = mb_counts.get('henkei_r_arm', '')
    target_ws["BS132"].alignment = Alignment(horizontal="center", vertical="center")
    target_ws["CE132"] = mb_counts.get('henkei_l_arm', '')
    target_ws["CE132"].alignment = Alignment(horizontal="center", vertical="center")
    target_ws["CQ132"] = mb_counts.get('henkei_r_leg', '')
    target_ws["CQ132"].alignment = Alignment(horizontal="center", vertical="center")
    target_ws["DC132"] = mb_counts.get('henkei_l_leg', '')
    target_ws["DC132"].alignment = Alignment(horizontal="center", vertical="center")
    if 'henkei' in m_fees and m_fees['henkei']:
        tf = m_fees['henkei']
        safe_apply_grid_row(target_ws, 136, 139, tf['price'], tf['count'], tf['total'], is_boxed=False)
    else:
        safe_apply_grid_row(target_ws, 136, 139, None, None, None, is_boxed=False)

    # 特別地域 (加算 Row 140..143)
    if 'tokubetsu' in m_fees:
        tf = m_fees['tokubetsu']
        safe_apply_grid_row(target_ws, 140, 143, tf['price'], tf['count'], tf['total'], is_boxed=False)

    # 往療料 (Row 144..147)
    if 'ouryou' in m_fees:
        tf = m_fees['ouryou']
        safe_apply_grid_row(target_ws, 144, 147, tf['price'], tf['count'], tf['total'], is_boxed=False)

    # 施術報告書交付料 (Row 148..151)
    prev_ym_mass = extract_report_prev_date(spot_ws, is_massage=True)
    target_ws["J148"] = f"施術報告書交付料　（前回支給：{prev_ym_mass}）"
    if 'houkoku' in m_fees:
        tf = m_fees['houkoku']
        safe_apply_grid_row(target_ws, 148, 151, tf['price'], tf['count'], tf['total'], is_boxed=False)
        
    # 明細書発行加算 (Row 152..155)
    if 'meisai' in m_fees and m_fees['meisai']:
        tf = m_fees['meisai']
        safe_apply_grid_row(target_ws, 152, 155, tf['price'], tf['count'], tf['total'], is_boxed=False)

    # 合計
    if 'goukei' in m_fees and m_fees['goukei'] is not None:
        target_ws["BF156"] = f"{format_currency_str(m_fees['goukei'])} 円"
    else:
        target_ws["BF156"] = None
        
    # 一部負担金 & 請求額
    target_ws["J160"] = extract_copayment_ratio_text(spot_ws, marks)
    if 'ichibu' in m_fees and m_fees['ichibu'] is not None:
        target_ws["BF160"] = f"{format_currency_str(m_fees['ichibu'])} 円"
    else:
        target_ws["BF160"] = None
        
    if 'seikyuu' in m_fees and m_fees['seikyuu'] is not None:
        target_ws["BF164"] = f"{format_currency_str(m_fees['seikyuu'])} 円"
    else:
        target_ws["BF164"] = None

    # 摘要欄 (DO87:EX167 を結合して全文表示)
    try:
        target_ws.merge_cells("DO87:EX167")
    except Exception:
        pass
    m_summary = mb_counts.get('summary') or spot_ws["BK72"].value
    if m_summary:
        target_ws["DO87"] = str(m_summary)
        target_ws["DO87"].alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
        target_ws["DO87"].font = Font(name="ＭＳ 明朝", size=9)

    # カレンダー (Row 172: 中央揃えで動的配置)
    cal_data = extract_calendar_marks_dynamic(spot_ws)
    for d, mark in cal_data.items():
        if 1 <= d <= 31:
            cal_col_idx = 30 + (d - 1) * 4
            cell = target_ws.cell(row=172, column=cal_col_idx, value=str(mark))
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(name="ＭＳ 明朝", size=11, bold=True)

    # 往療又は訪問の理由 (動的複数選択対応)
    target_ws["J177"] = detect_visit_reasons(spot_ws, img_coords)

    # 施術証明欄 (完全動的抽出)
    c_sec = extract_cert_section_dynamic(spot_ws)
    target_ws["DC181"] = detect_practitioner_location_type(spot_ws, img_coords)
    if c_sec.get('zip'): target_ws["CR185"] = f"〒{str(c_sec['zip']).replace('〒', '')}"
    if c_sec.get('date'): target_ws["M188"] = str(c_sec['date'])
    if c_sec.get('addr'): target_ws["CR188"] = str(c_sec['addr'])
    if c_sec.get('reg_no'): target_ws["M195"] = str(c_sec['reg_no'])
    
    try:
        target_ws.merge_cells("CR192:EX195")
    except Exception:
        pass
    if c_sec.get('clinic_name'):
        target_ws["CR192"] = str(c_sec['clinic_name'])
        target_ws["CR192"].alignment = Alignment(vertical="center", horizontal="left")
        target_ws["CR192"].font = Font(name="ＭＳ 明朝", size=10)
        
    if c_sec.get('manager_name'): target_ws["CR196"] = str(c_sec['manager_name'])
    if c_sec.get('tel'): target_ws["EG196"] = str(c_sec['tel'])

    # 申請欄 (完全動的抽出)
    a_sec = extract_applicant_section_dynamic(spot_ws)
    if a_sec.get('zip'): target_ws["CR200"] = f"〒{str(a_sec['zip']).replace('〒', '')}"
    if a_sec.get('date'): target_ws["M204"] = str(a_sec['date'])
    if a_sec.get('addr'): target_ws["CR204"] = str(a_sec['addr'])
    if a_sec.get('kouiki'): target_ws["M208"] = str(a_sec['kouiki'])
    if a_sec.get('name'): target_ws["CR211"] = str(a_sec['name'])
    if a_sec.get('tel'): target_ws["EG211"] = str(a_sec['tel'])

    # 支払機関欄 (完全動的判定)
    pay_sec = detect_payment_section(spot_ws, img_coords)
    
    target_ws["L218"] = "①．" if pay_sec['pay_1'] else "1．"
    target_ws["AH218"] = "②．" if pay_sec['pay_2'] else "2．"
    target_ws["L221"] = "③．" if pay_sec['pay_3'] else "3．"
    target_ws["AH221"] = "④．" if pay_sec['pay_4'] else "4．"
    
    target_ws["BD218"] = "①.　普通" if pay_sec['dep_1'] else "1.　普通"
    target_ws["BS218"] = "②.　当座" if pay_sec['dep_2'] else "2.　当座"
    target_ws["BD221"] = "③.　通知" if pay_sec['dep_3'] else "3.　通知"
    target_ws["BS221"] = "④.　別段" if pay_sec['dep_4'] else "4.　別段"
    
    # 金融機関名（右側の入力枠 CW215:DQ223 に中央揃えで配置、長い名前も自動文字縮小対応）
    b_name = pay_sec['bank_name']
    try:
        target_ws.merge_cells("CW215:DQ223")
    except Exception:
        pass
    target_ws["CW215"] = b_name
    b_font_size = 7.5 if len(b_name) > 10 else (8.5 if len(b_name) > 6 else 10.0)
    target_ws["CW215"].font = Font(name="ＭＳ 明朝", size=b_font_size)
    target_ws["CW215"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    b_t = pay_sec['bank_type']
    target_ws["DR215"] = "○銀行" if "銀行" in b_t else "銀行"
    target_ws["DR218"] = "○金庫" if any(k in b_t for k in ["金庫", "信金"]) else "金庫"
    target_ws["DR221"] = "○農協" if any(k in b_t for k in ["農協", "JA"]) else "農協"
    
    # 支店名（DY215:EN223 に配置、長い名前も自動文字縮小対応）
    br_name = pay_sec['branch_name']
    try:
        target_ws.merge_cells("DY215:EN223")
    except Exception:
        pass
    target_ws["DY215"] = br_name
    br_font_size = 7.5 if len(br_name) > 10 else (8.5 if len(br_name) > 6 else 10.0)
    target_ws["DY215"].font = Font(name="ＭＳ 明朝", size=br_font_size)
    target_ws["DY215"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    br_t = pay_sec['branch_type']
    target_ws["EO215"] = "○本店" if "本店" in br_t else "本店"
    target_ws["EO218"] = "○支店" if "支店" in br_t else "支店"
    target_ws["EO221"] = "○出張所" if "出張所" in br_t else "出張所"
    
    # 口座番号 & 口座名義（完全動的抽出: 右詰めで配置）
    acc_boxes = ["CG224", "CL224", "CQ224", "CV224", "DA224", "DF224", "DK224", "DP224"]
    for col in acc_boxes:
        target_ws[col] = None
        
    p_acc = extract_payment_account_info_dynamic(spot_ws)
    acc_digits = p_acc['digits']
    if acc_digits:
        digs = acc_digits[-8:] if len(acc_digits) > 8 else acc_digits
        start_idx = len(acc_boxes) - len(digs)
        for i, d in enumerate(digs):
            target_ws[acc_boxes[start_idx + i]] = str(d)
        
    if p_acc.get('holder'):
        target_ws["AD224"] = str(p_acc['holder'])

    # 同意記録 (動的探索)
    c_rec = extract_consent_record(spot_ws)
    if c_rec:
        target_ws["J233"] = c_rec['doc_name']
        target_ws["AJ233"] = c_rec['address']
        target_ws["CA233"] = c_rec['consent_date']
        target_ws["DE233"] = c_rec['disease']
        target_ws["EF233"] = c_rec['period']
    else:
        target_ws["J233"] = None
        target_ws["AJ233"] = None
        target_ws["DE233"] = None
        target_ws["EF233"] = None

    # 委任状欄 (完全動的抽出)
    d_sec = extract_delegation_section_dynamic(spot_ws)
    if d_sec.get('date'): target_ws["CP243"] = str(d_sec['date'])
    
    target_ws["J247"] = "申請者"
    target_ws["J247"].alignment = Alignment(horizontal="center", vertical="center")
    
    target_ws["CD247"] = "代理人"
    target_ws["CD247"].alignment = Alignment(horizontal="center", vertical="center")
    
    target_ws["J255"] = "（被保険者）"
    target_ws["J255"].alignment = Alignment(horizontal="center", vertical="center")
    
    target_ws["CD255"] = None
    
    applicant_addr = str(d_sec.get('applicant_addr') or "")
    target_ws["AA247"] = f"住所　{applicant_addr}" if applicant_addr else "住所　"
    target_ws["AA247"].alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
    target_ws["AA247"].font = Font(name="ＭＳ 明朝", size=10)
    
    target_ws["CU247"] = format_proxy_address(d_sec.get('delegate_addr'))
    target_ws["CU247"].alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
    target_ws["CU247"].font = Font(name="ＭＳ 明朝", size=8.5)
    
    applicant_name = str(d_sec.get('applicant_name') or "")
    target_ws["AA255"] = f"氏名　{applicant_name}" if applicant_name else "氏名　"
    target_ws["AA255"].alignment = Alignment(vertical="center", horizontal="left")
    target_ws["AA255"].font = Font(name="ＭＳ 明朝", size=10)
    
    proxy_name = str(d_sec.get('delegate_name') or "")
    target_ws["CU255"] = f"氏名　{proxy_name}" if proxy_name else "氏名　"
    target_ws["CU255"].alignment = Alignment(vertical="center", horizontal="left")
    target_ws["CU255"].font = Font(name="ＭＳ 明朝", size=10)

    patient_name = h_data.get('name') or spot_ws["O40"].value or spot_ws.title
    return str(patient_name).strip()


# ==========================================
# Main Streamlit App
# ==========================================

def main():
    if 'uploader_key' not in st.session_state:
        st.session_state.uploader_key = 0

    st.markdown('<div class="main-header">spotlog ➔ 基準様式 Excel自動変換ツール</div>', unsafe_allow_html=True)
    st.markdown('<div class="sub-header">spotlog の Excelファイル（.xlsx）から<br>基準様式フォーマットへ一括変換します</div>', unsafe_allow_html=True)

    base_dir = os.path.dirname(os.path.abspath(__file__))
    tpl_hari_path = find_template_file("鍼灸", base_dir)
    tpl_massage_path = find_template_file("マッサージ", base_dir)
    
    has_templates = tpl_hari_path and tpl_massage_path
    
    if not has_templates:
        st.warning("⚠️ 基準様式テンプレートファイルが見つかりません。")
        return

    # 前回変換結果の表示エリア（新しくファイルが選択されたら自動クリア）
    if 'last_result' not in st.session_state:
        st.session_state.last_result = None

    header_html = """
    <div class="dropzone-header-flex">
        <div class="dropzone-title-text"><span class="material-symbols-outlined" style="margin-right: 0.4rem; font-size: 1.25rem !important; color: #475569;">folder_open</span>変換する spotlog ファイルを選択またはドロップ（複数可）</div>
        <details class="spec-popover">
            <summary class="spec-info-btn" title="機能と変換仕様を表示">i</summary>
            <div class="spec-popover-card">
                <h4><span class="material-symbols-outlined" style="margin-right: 0.35rem; font-size: 1.15rem !important; vertical-align: -2px; color: #475569;">push_pin</span>機能と変換仕様</h4>
                <ul>
                    <li><strong>対応ファイル形式:</strong> .xlsx 形式のExcelファイルに対応しています（複数ファイル可）。</li>
                    <li><strong>spotlog様式の自動判定:</strong> 右上の様式表記（様式第５号の３＝はり・きゅう、様式第５号の４＝マッサージ）を厳密に照合します。spotlog形式以外のシートや無関係なファイルは自動検知され、安全にスキップされます。</li>
                    <li><strong>自動転記される主な項目:</strong> 患者情報（氏名・住所・生年月日）、公費・保険情報、傷病名、往療理由、支払区分、口座情報、金額計算欄（訪問施術料・電療加算・温罨法加算・施術報告書交付料等）など、spotlogに記載された各項目を基準様式の所定の欄へ自動で正確に転記します。</li>
                    <li><strong>基準様式対象外（逓減・算定項目等）の自動スキップ:</strong> 基準様式（様式第5号の1/2）に記載枠が存在しない「月16回以降の50%逓減」「訪問施術料4（10〜19人）」「訪問施術料5（20人以上）」「施設集中率80%逓減」が算定されているシートは、変換対象外として自動検知しスキップ・通知します。</li>
                    <li><strong>日時付き個別ファイル保存:</strong> 実行ごとに <code>基準様式_一括変換結果_YYYYMMDD_HHMMSS.xlsx</code> という日時付きの名前で新規保存されるため、過去の変換ファイルが上書きされて消えることはありません。</li>
                </ul>
            </div>
        </details>
    </div>
    """
    st.markdown(header_html, unsafe_allow_html=True)

    uploaded_files = st.file_uploader(
        "spotlog_uploader",
        type=["xlsx"],
        accept_multiple_files=True,
        key=f"uploader_{st.session_state.uploader_key}",
        label_visibility="collapsed"
    )

    if uploaded_files:
        # ファイル設置時：DropZone（#0A84FF / 16px角丸）完全一致スタイル
        st.markdown("""
            <style>
            div[data-testid="stFileUploader"] section,
            div[data-testid="stFileUploaderDropzone"],
            section[data-testid="stFileUploadDropzone"],
            [data-testid="stFileUploader"] > section,
            [data-testid="stFileUploadDropzone"] {
                background: #EFF6FF !important;
                background-color: #EFF6FF !important;
                border: 2px solid #0A84FF !important;
                border-style: solid !important;
                border-color: #0A84FF !important;
                border-radius: 16px !important;
                padding: 3.6rem 2rem !important;
                min-height: 250px !important;
                margin-bottom: 0 !important;
                box-shadow: 0 4px 16px rgba(10, 132, 255, 0.15) !important;
            }
            /* ファイル設置時のホバー：#3090FF (STYLE_HOVER) */
            div[data-testid="stFileUploader"] section:hover,
            div[data-testid="stFileUploaderDropzone"]:hover,
            section[data-testid="stFileUploadDropzone"]:hover,
            [data-testid="stFileUploadDropzone"]:hover {
                background: #F4F8FF !important;
                background-color: #F4F8FF !important;
                border: 2px solid #3090FF !important;
                border-color: #3090FF !important;
            }
            div[data-testid="stFileUploaderFile"] {
                display: flex !important;
                align-items: center !important;
                justify-content: space-between !important;
                background-color: #FFFFFF !important;
                border: 1px solid #BFDBFE !important;
                border-radius: 10px !important;
                padding: 0.6rem 1rem !important;
                margin-bottom: 0.5rem !important;
                color: #1E293B !important;
                box-shadow: 0 2px 8px rgba(0, 0, 0, 0.05) !important;
            }
            [data-testid="stFileUploaderFile"] svg {
                fill: #0A84FF !important;
                color: #0A84FF !important;
            }
            [data-testid="stFileUploaderFile"] span,
            [data-testid="stFileUploaderFile"] p {
                color: #1E293B !important;
                font-weight: 500 !important;
            }
            [data-testid="stFileUploaderFile"] small {
                color: #64748B !important;
            }
            /* ファイル取り消し（削除）ボタン：コードと同じ #FF453A ホバー */
            [data-testid="stFileUploaderDeleteFile"] {
                display: flex !important;
                align-items: center !important;
                justify-content: center !important;
                margin-left: auto !important;
            }
            [data-testid="stFileUploaderDeleteFile"] button {
                display: flex !important;
                align-items: center !important;
                justify-content: center !important;
                width: 28px !important;
                height: 28px !important;
                border-radius: 50% !important;
                border: none !important;
                background-color: #F1F5F9 !important;
                color: #64748B !important;
                cursor: pointer !important;
                transition: all 0.2s ease !important;
            }
            [data-testid="stFileUploaderDeleteFile"] button:hover {
                background-color: #FFEBEA !important;
                color: #FF453A !important;
            }
            [data-testid="stFileUploaderDeleteFile"] svg {
                display: block !important;
                width: 14px !important;
                height: 14px !important;
                fill: currentColor !important;
            }
            </style>
        """, unsafe_allow_html=True)
        
        # 新しいファイルがアップロードされたら前回の結果表示をクリア
        st.session_state.last_result = None
        
        btn_placeholder = st.empty()
        start_btn = btn_placeholder.button(":material/play_arrow: 基準様式へ一括変換開始", use_container_width=True, type="primary")
        
        if start_btn:
            btn_placeholder.empty()  # 押したらボタンを消去して進行状況バーを上に詰める
            
            # ボタンを押した瞬間に0%のプログレスバーを即座に表示！
            progress_bar = st.progress(0.0)
            status_text = st.empty()
            status_text.markdown('<span class="material-symbols-outlined" style="margin-right: 0.3rem; vertical-align: -3px; color: #475569;">hourglass_top</span>**準備中... 0%** — ファイルを読み込んでいます...', unsafe_allow_html=True)
            
            import zipfile
            all_sheets_data = []
            for file_idx, file_item in enumerate(uploaded_files):
                try:
                    bytes_data = file_item.getvalue()
                    if file_item.name.lower().endswith('.zip'):
                        with zipfile.ZipFile(io.BytesIO(bytes_data)) as z:
                            for zname in z.namelist():
                                if zname.lower().endswith('.xlsx') and not zname.startswith('__MACOSX/'):
                                    zbytes = z.read(zname)
                                    wb_in = openpyxl.load_workbook(filename=io.BytesIO(zbytes), read_only=True)
                                    for sheetname in wb_in.sheetnames:
                                        all_sheets_data.append({
                                            'file_name': f"{file_item.name} > {zname}",
                                            'bytes_data': zbytes,
                                            'sheet_name': sheetname
                                        })
                                    wb_in.close()
                    else:
                        wb_in = openpyxl.load_workbook(filename=io.BytesIO(bytes_data), read_only=True)
                        for sheetname in wb_in.sheetnames:
                            all_sheets_data.append({
                                'file_name': file_item.name,
                                'bytes_data': bytes_data,
                                'sheet_name': sheetname
                            })
                        wb_in.close()
                except Exception as e:
                    st.warning(f"⚠️ ファイル `{file_item.name}` の読み込みに失敗しました: {e}")
                    
            total_sheets = len(all_sheets_data)
            
            if total_sheets == 0:
                st.error("処理可能なシートが見つかりませんでした。")
                return
            
            used_sheet_names = set()
            count_hari = 0
            count_massage = 0
            skipped_sheets = []
            
            loaded_workbooks = {}

            tpl_hari_wb = openpyxl.load_workbook(tpl_hari_path)
            tpl_mass_wb = openpyxl.load_workbook(tpl_massage_path)

            master_wb = openpyxl.load_workbook(tpl_hari_path)
            tpl_hari_sheet = master_wb.active
            tpl_hari_sheet.title = "__TPL_HARI__"
            
            tpl_mass_sheet = clone_worksheet_to_wb(tpl_mass_wb.active, master_wb, "__TPL_MASS__")

            for idx, item in enumerate(all_sheets_data, start=1):
                ratio = idx / total_sheets
                percent = int(ratio * 100)
                progress_bar.progress(ratio)
                status_text.markdown(f'<span class="material-symbols-outlined" style="margin-right: 0.3rem; vertical-align: -3px; color: #475569;">sync</span>**変換処理中... {percent}% ({idx}/{total_sheets}件)** — 処理中: 【{item["file_name"]}】 シート「{item["sheet_name"]}」', unsafe_allow_html=True)
                
                fname = item['file_name']
                if fname not in loaded_workbooks:
                    loaded_workbooks[fname] = openpyxl.load_workbook(filename=io.BytesIO(item['bytes_data']), data_only=False)
                
                wb_source = loaded_workbooks[fname]
                source_ws = wb_source[item['sheet_name']]
                
                # 右上の様式番号（様式第５号の３ / 様式第５号の４）を厳密チェック
                sheet_type = detect_spotlog_type(source_ws)
                
                if sheet_type is None:
                    skipped_sheets.append(f"【{item['file_name']}】 シート「{item['sheet_name']}」 (非spotlog様式)")
                    continue
                
                # 基準様式対象外の厳密チェック（訪問施術料4・5、月16回以降の50%逓減、集中率80%逓減など）
                is_ineligible, skip_reason = check_is_ineligible_for_standard_form(source_ws, sheet_type)
                if is_ineligible:
                    patient_p = source_ws["O40"].value or item['sheet_name']
                    skipped_sheets.append(f"【{item['file_name']}】 シート「{item['sheet_name']}」 ({patient_p} 様: {skip_reason}基準様式対象外)")
                    continue
                
                if sheet_type == "マッサージ":
                    converted_ws = master_wb.copy_worksheet(tpl_mass_sheet)
                    patient_name = convert_massage_dynamic(source_ws, converted_ws)
                    base_tab = f"{patient_name}_マ"
                    count_massage += 1
                else:
                    converted_ws = master_wb.copy_worksheet(tpl_hari_sheet)
                    patient_name = convert_acupuncture_dynamic(source_ws, converted_ws)
                    base_tab = f"{patient_name}_はり"
                    count_hari += 1
                    
                # タブ色を消去（デフォルトの白/透明に設定）
                converted_ws.sheet_properties.tabColor = None
                
                clean_name = sanitize_sheet_name(base_tab)
                if clean_name in used_sheet_names:
                    counter = 2
                    while sanitize_sheet_name(f"{base_tab}_{counter}") in used_sheet_names:
                        counter += 1
                    clean_name = sanitize_sheet_name(f"{base_tab}_{counter}")
                used_sheet_names.add(clean_name)
                
                converted_ws.title = clean_name

            # テンプレート用の一時シートを削除
            if "__TPL_HARI__" in master_wb.sheetnames:
                master_wb.remove(tpl_hari_sheet)
            if "__TPL_MASS__" in master_wb.sheetnames:
                master_wb.remove(tpl_mass_sheet)

            total_converted = count_hari + count_massage
            
            if total_converted == 0:
                st.session_state.last_result = {
                    'error_msg': '変換可能な spotlog シートが見つかりませんでした。'
                }
                st.session_state.uploader_key += 1
                st.rerun()

            # 日時付きファイル名を生成（日本時間 JST: UTC+9 で固定）
            jst = timezone(timedelta(hours=9))
            timestamp_str = datetime.now(jst).strftime("%Y%m%d_%H%M%S")
            output_filename = f"基準様式_一括変換結果_{timestamp_str}.xlsx"
            
            # メモリバッファに書き出し（Webダウンロード用）
            out_buf = io.BytesIO()
            master_wb.save(out_buf)
            excel_bytes = out_buf.getvalue()

            # ポップアップ通知（Toast）を表示
            st.toast(f"変換完了: {output_filename}", icon="✅")

            # 完了結果を保持し、ファイル選択エリアのみをリセット
            st.session_state.last_result = {
                'output_filename': output_filename,
                'total_converted': total_converted,
                'count_hari': count_hari,
                'count_massage': count_massage,
                'skipped_sheets': skipped_sheets,
                'excel_bytes': excel_bytes,
                'should_auto_download': True
            }
            st.session_state.uploader_key += 1
            st.rerun()

    # 変換完了結果の常時表示（ファイルエリアのみリセット後も消えずに残る）
    elif st.session_state.get('last_result'):
        res = st.session_state.last_result
        
        if 'error_msg' in res:
            st.markdown(f'<div class="error-summary-card"><span class="material-symbols-outlined" style="font-size: 1.35rem !important; color: #DC2626;">error</span><span class="error-summary-title">{res["error_msg"]}</span></div>', unsafe_allow_html=True)
        else:
            # 自動ダウンロード発動用（初回のみ発動）
            if res.get('should_auto_download') and res.get('excel_bytes'):
                res['should_auto_download'] = False
                b64_data = base64.b64encode(res['excel_bytes']).decode()
                auto_dl_html = f"""
                    <a id="auto_dl_anchor" href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64_data}" download="{res['output_filename']}" style="display:none;"></a>
                    <script>
                        setTimeout(function() {{
                            var link = document.getElementById('auto_dl_anchor');
                            if (link) {{ link.click(); }}
                        }}, 100);
                    </script>
                """
                st.components.v1.html(auto_dl_html, height=0)

            skipped_html = ""
            if res.get('skipped_sheets'):
                items_html = "".join([f"<li>{s}</li>" for s in res['skipped_sheets']])
                skipped_html = f'<div class="result-skipped-section"><div class="result-skipped-title"><span class="material-symbols-outlined" style="font-size: 1.15rem !important; margin-right: 0.35rem; vertical-align: -2px; color: #64748B;">fast_forward</span>基準様式対象外等によりスキップされたシート ({len(res["skipped_sheets"])}件):</div><div class="result-skipped-scroll"><ul>{items_html}</ul></div></div>'

            result_card_html = f'<div class="result-summary-card"><div class="result-header"><span class="material-symbols-outlined" style="font-size: 1.75rem !important; color: #1E293B;">check_circle</span><div class="result-title-group"><div class="result-main-title">変換が正常に完了しました</div><div class="result-sub-title">お使いの端末へ自動ダウンロードされます</div></div></div><div class="result-details-grid"><div class="result-detail-row"><span class="result-label"><span class="material-symbols-outlined" style="font-size: 1.1rem !important; margin-right: 0.3rem; vertical-align: -2px; color: #475569;">analytics</span>変換件数:</span><span class="result-value"><strong>{res["total_converted"]} 件</strong>（はり・きゅう: {res["count_hari"]}件 / マッサージ: {res["count_massage"]}件）</span></div><div class="result-detail-row"><span class="result-label"><span class="material-symbols-outlined" style="font-size: 1.1rem !important; margin-right: 0.3rem; vertical-align: -2px; color: #475569;">description</span>保存ファイル:</span><span class="result-value file-path-badge">{res["output_filename"]}</span></div></div>{skipped_html}</div>'
            st.markdown(result_card_html, unsafe_allow_html=True)

if __name__ == "__main__":
    main()
